import datetime
import io
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.analyzer import COL_PAD, ROW_PAD, _apply_sheet_style, sort_dates
from app.drive_uploader import download_drive_file, list_drive_files
from app.logger import get_logger

logger = get_logger("merger")

_CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "education_config.json")


def _load_education_config():
    """education_config.json을 로드한다. 파일 없으면 빈 설정 반환."""
    try:
        with open(_CONFIG_PATH, encoding="utf-8") as f:
            config = json.load(f)
        logger.info("교육국 설정 로드: nt_members=%s, excluded=%s",
                     config.get("nt_members", []), config.get("excluded_members", []))
        return config
    except (FileNotFoundError, json.JSONDecodeError) as exc:
        logger.warning("교육국 설정 파일 로드 실패 (%s) — 기본값 사용", exc)
        return {"nt_members": [], "excluded_members": []}


def _normalize_room_name(room):
    """방이름에서 '꿀성경' 접두사와 구분자를 제거하여 정규화한다."""
    for prefix in ("꿀성경 - ", "꿀성경 ", "꿀성경"):
        if room.startswith(prefix):
            room = room[len(prefix):]
            break
    return room.strip(" -_")


_FILENAME_DATE_RE = re.compile(r"_(\d{8})_(\d{4})[_.]")


def _extract_date_from_filename(name):
    """파일명에서 YYYYMMDD_HHMM 형식의 날짜시간을 추출한다."""
    if not name:
        return None
    m = _FILENAME_DATE_RE.search(name)
    if m:
        return f"{m.group(1)}_{m.group(2)}"
    return None


def _extract_room_from_filename(name):
    """파일명 끝에서 방이름을 추출한다.

    패턴: 꿀성경_방장_YYYYMMDD_HHMM_방이름.xlsx → 방이름
    """
    if not name:
        return name
    # 확장자 제거
    stem = name.rsplit(".", 1)[0] if "." in name else name
    # 꿀성경_방장_날짜_시간_방이름 패턴: 최소 5개 부분
    parts = stem.split("_")
    if len(parts) >= 5 and parts[0] == "꿀성경":
        # 날짜 부분(YYYYMMDD)과 시간 부분(HHMM) 확인
        date_part = parts[2]
        time_part = parts[3]
        if re.match(r"^\d{8}$", date_part) and re.match(r"^\d{4}$", time_part):
            # 4번째 인덱스부터 끝까지가 방이름 (_로 재결합)
            room = "_".join(parts[4:])
            return _normalize_room_name(room)
    return name


def select_latest_per_room(files):
    """파일 목록에서 방별 최신 파일을 선택한다.

    Args:
        files: [{"id": str, "name": str, "modifiedTime": str}, ...]

    Returns:
        list: 방별 최신 파일만 포함된 리스트
    """
    rooms = {}
    for f in files:
        room = _extract_room_from_filename(f["name"])
        existing = rooms.get(room)
        if existing is None or f["modifiedTime"] > existing["modifiedTime"]:
            rooms[room] = f
    result = list(rooms.values())
    logger.info("방별 최신 파일 선택: %d개 방 → %d개 파일", len(rooms), len(result))
    return result


def read_meta_from_xlsx(xlsx_bytes):
    """XLSX 바이트에서 _메타 시트를 읽어 dict로 반환한다. 없으면 None."""
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        if "_메타" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["_메타"]
        meta = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] is not None:
                meta[str(row[0])] = str(row[1]) if row[1] is not None else ""
        wb.close()
        return meta
    except Exception as exc:
        logger.warning("메타데이터 읽기 실패: %s", exc)
        return None


def _find_header_row(ws):
    """'이름' 컬럼을 찾아 (rows_iter, trimmed_header, col_offset) 반환."""
    rows_iter = ws.iter_rows(values_only=True)
    for row in rows_iter:
        values = list(row)
        for i, v in enumerate(values):
            if v == "이름":
                return rows_iter, values[i:], i
    return iter([]), None, 0


def read_users_from_xlsx(xlsx_bytes, track_mode):
    """XLSX 바이트에서 사용자 데이터를 추출한다.

    Args:
        xlsx_bytes: XLSX 파일 바이트
        track_mode: "single" 또는 "dual"

    Returns:
        dict: single → {user: {"dates": set, "emoji": str}}
              dual → {user: {"dates_old": set, "dates_new": set, "emoji": str}}
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    users = {}

    if track_mode == "dual":
        for sheet_name, date_key in [("구약 진도표", "dates_old"), ("신약 진도표", "dates_new")]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows_iter, header, col_offset = _find_header_row(ws)
            if not header:
                continue
            # 날짜 컬럼: "이름", "이모티콘" 뒤부터
            date_cols = list(header[2:])
            for row in rows_iter:
                if not row or col_offset >= len(row) or not row[col_offset]:
                    continue
                name = str(row[col_offset])
                emoji = str(row[col_offset + 1]) if (col_offset + 1) < len(row) and row[col_offset + 1] else ""
                entry = users.setdefault(name, {"dates_old": set(), "dates_new": set(), "emoji": emoji})
                if not entry["emoji"] and emoji:
                    entry["emoji"] = emoji
                for i, date_val in enumerate(date_cols):
                    cell_idx = col_offset + 2 + i
                    cell_val = row[cell_idx] if cell_idx < len(row) else None
                    if cell_val == "O" and date_val:
                        entry[date_key].add(str(date_val))
    else:
        sheet_name = "꿀성경 진도표"
        if sheet_name not in wb.sheetnames:
            # 첫 번째 시트 사용 (호환성)
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows_iter, header, col_offset = _find_header_row(ws)
        if header:
            date_cols = list(header[2:])
            for row in rows_iter:
                if not row or col_offset >= len(row) or not row[col_offset]:
                    continue
                name = str(row[col_offset])
                emoji = str(row[col_offset + 1]) if (col_offset + 1) < len(row) and row[col_offset + 1] else ""
                entry = users.setdefault(name, {"dates": set(), "emoji": emoji})
                if not entry["emoji"] and emoji:
                    entry["emoji"] = emoji
                for i, date_val in enumerate(date_cols):
                    cell_idx = col_offset + 2 + i
                    cell_val = row[cell_idx] if cell_idx < len(row) else None
                    if cell_val == "O" and date_val:
                        entry["dates"].add(str(date_val))

    wb.close()
    return users


def _classify_education_users(users, config):
    """교육국 사용자를 설정 파일 기반으로 성경일독/신약일독으로 분류한다.

    Args:
        users: {user: {"dates": set, "emoji": str}}
        config: {"nt_members": [...], "excluded_members": [...]}

    Returns:
        dict: {"bible": {user: data}, "nt": {user: data}}
    """
    nt_keywords = config.get("nt_members", [])
    excluded_keywords = config.get("excluded_members", [])

    def _matches(user, keywords):
        return any(kw in user for kw in keywords)

    result = {"bible": {}, "nt": {}}
    for user, data in users.items():
        if _matches(user, excluded_keywords):
            logger.info("교육국 미참여 제외: %s", user)
            continue
        if _matches(user, nt_keywords):
            result["nt"][user] = {"dates": data["dates"].copy(), "emoji": data["emoji"]}
        else:
            result["bible"][user] = {"dates": data["dates"].copy(), "emoji": data["emoji"]}

    excluded_count = sum(1 for u in users if _matches(u, excluded_keywords))
    logger.info("교육국 분류: 성경일독 %d명, 신약일독 %d명, 제외 %d명",
                len(result["bible"]), len(result["nt"]), excluded_count)
    return result


def _merge_user_into(target, user, dates, emoji, leader):
    """대상 dict에 사용자 날짜를 합집합으로 병합한다."""
    if user in target:
        target[user]["dates"].update(dates)
        # 담당은 교육국방이 아닌 쪽 우선
        if leader and not target[user].get("leader"):
            target[user]["leader"] = leader
    else:
        target[user] = {
            "dates": set(dates),
            "emoji": emoji,
            "leader": leader or "",
        }


def _merge_dual_user_into(target, user, dates_old, dates_new, emoji, leader):
    """투트랙 사용자를 구약/신약 분리하여 병합한다."""
    if user in target:
        target[user]["dates_old"].update(dates_old)
        target[user]["dates_new"].update(dates_new)
        if leader and not target[user].get("leader"):
            target[user]["leader"] = leader
    else:
        target[user] = {
            "dates_old": set(dates_old),
            "dates_new": set(dates_new),
            "emoji": emoji,
            "leader": leader or "",
        }


def _is_saturday(md_str):
    """'M/D' 문자열이 2026년 기준 토요일인지 판별한다."""
    month, day = md_str.split("/")
    return datetime.date(2026, int(month), int(day)).weekday() == 5


def _compute_stats(users, all_dates_sorted):
    """성경일독/신약일독 시트용 통계 문자열을 생성한다."""
    all_dates_set = set(all_dates_sorted)
    num_dates = len(all_dates_sorted)
    members = [u for u in users if users[u]["dates"]]
    num_members = len(members)
    num_perfect = sum(1 for u in members if users[u]["dates"] >= all_dates_set)
    rate = (num_perfect / num_members * 100) if num_members else 0
    return f"진행: {num_dates}일 | 참여: {num_members}명 | 완독: {num_perfect}명 ({rate:.0f}%)"


def _compute_dual_stats(dual_users, all_dates_sorted):
    """투트랙 시트용 통계 문자열을 생성한다.

    완독 기준: 구약(토요일 제외) AND 신약 모두 완독.
    """
    num_dates = len(all_dates_sorted)
    num_members = len(dual_users)
    all_old = set()
    all_new = set()
    for data in dual_users.values():
        all_old.update(data["dates_old"])
        all_new.update(data["dates_new"])
    old_expected = {d for d in all_old if not _is_saturday(d)}
    new_expected = all_new
    num_perfect = sum(
        1 for data in dual_users.values()
        if data["dates_old"] >= old_expected and data["dates_new"] >= new_expected
    )
    rate = (num_perfect / num_members * 100) if num_members else 0
    return f"진행: {num_dates}일 | 참여: {num_members}명 | 완독: {num_perfect}명 ({rate:.0f}%)"


def _insert_stats_row(ws, stats_text, num_headers):
    """타이틀과 헤더 사이에 통계 부제 행을 삽입한다."""
    stats_row = 2 + ROW_PAD  # row 3 위치에 삽입 → 헤더가 row 4로 이동

    # openpyxl의 insert_rows는 merged cell range를 올바르게 시프트하지 않으므로
    # 삽입 지점 이후의 병합을 수동으로 처리한다.
    ranges_to_shift = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= stats_row:
            ranges_to_shift.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
            ws.unmerge_cells(str(rng))

    ws.insert_rows(stats_row, 1)

    # 시프트된 위치로 재병합
    for min_r, min_c, max_r, max_c in ranges_to_shift:
        ws.merge_cells(start_row=min_r + 1, start_column=min_c,
                       end_row=max_r + 1, end_column=max_c)

    last_col = num_headers + COL_PAD
    ws.merge_cells(start_row=stats_row, start_column=1 + COL_PAD,
                   end_row=stats_row, end_column=last_col)
    medium = Side(style="medium")
    stats_border = Border(top=medium, bottom=medium, left=medium, right=medium)
    stats_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell = ws.cell(row=stats_row, column=1 + COL_PAD, value=stats_text)
    cell.font = Font(name="맑은 고딕", size=13, bold=True)
    cell.fill = stats_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = stats_border
    for c in range(1 + COL_PAD + 1, last_col + 1):
        mc = ws.cell(row=stats_row, column=c)
        mc.fill = stats_fill
        mc.border = stats_border
    ws.row_dimensions[stats_row].height = 28
    ws.freeze_panes = "B5"  # 기존 "B4" → 1행 추가로 "B5"


def merge_files(dual_mode="separate"):
    """Drive에서 파일을 가져와 통합한다.

    Args:
        dual_mode: "split" — 투트랙 인원을 성경일독/신약일독에 분산 (기존 방식)
                   "separate" — 투트랙 인원을 별도 시트로 분리 (기본값)

    Returns:
        dict: 성공 시 {"success": True, "bible_users": dict, "nt_users": dict,
                      "dual_users": dict, "processed_rooms": list, "skipped_files": list}
              실패 시 {"success": False, "message": str}
    """
    # 1. Drive 파일 목록 조회
    list_result = list_drive_files()
    if not list_result["success"]:
        return list_result

    all_files = [f for f in list_result["files"] if "통합" not in f.get("name", "")]
    if not all_files:
        return {"success": False, "message": "Drive 폴더에 꿀성경 파일이 없습니다."}

    # 2. 방별 최신 파일 선택
    latest_files = select_latest_per_room(all_files)
    logger.info("통합 대상 파일: %d개", len(latest_files))

    # 3. 교육국 설정 로드
    edu_config = _load_education_config()

    bible_users = {}
    nt_users = {}
    dual_users = {}
    processed_rooms = []
    skipped_files = []
    oldest_file_date = None

    # 4. 파일 병렬 다운로드
    downloads = {}
    with ThreadPoolExecutor(max_workers=min(len(latest_files), 8)) as pool:
        futures = {
            pool.submit(download_drive_file, f["id"]): f for f in latest_files
        }
        for future in as_completed(futures):
            file_info = futures[future]
            downloads[file_info["id"]] = future.result()

    # 5. 파일별 처리
    for file_info in latest_files:
        file_id = file_info["id"]
        file_name = file_info["name"]

        dl_result = downloads[file_id]
        if not dl_result["success"]:
            skipped_files.append({"name": file_name, "reason": dl_result["message"]})
            continue

        xlsx_bytes = dl_result["data"]

        # 메타데이터 읽기
        meta = read_meta_from_xlsx(xlsx_bytes)
        if meta is None:
            skipped_files.append({"name": file_name, "reason": "메타데이터 없음 — 재업로드 필요"})
            continue

        schedule_type = meta.get("schedule_type", "unknown")
        track_mode = meta.get("track_mode", "single")
        leader = meta.get("leader", "")
        room_name = meta.get("room_name", "")

        logger.info("파일 처리: %s (schedule=%s, track=%s, leader=%s)",
                     file_name, schedule_type, track_mode, leader)

        # 사용자 데이터 추출
        users = read_users_from_xlsx(xlsx_bytes, track_mode)

        if schedule_type == "bible":
            for user, data in users.items():
                _merge_user_into(bible_users, user, data["dates"], data["emoji"], leader)
        elif schedule_type == "nt":
            for user, data in users.items():
                _merge_user_into(nt_users, user, data["dates"], data["emoji"], leader)
        elif schedule_type == "dual":
            dual_excluded = edu_config.get("dual_excluded_members", [])
            if dual_mode == "separate":
                for user, data in users.items():
                    if any(kw in user for kw in dual_excluded):
                        logger.info("투트랙 제외: %s", user)
                        continue
                    dates_old = data.get("dates_old", set())
                    dates_new = data.get("dates_new", set())
                    if dates_old or dates_new:
                        _merge_dual_user_into(dual_users, user, dates_old, dates_new, data["emoji"], leader)
            else:
                for user, data in users.items():
                    if data.get("dates_old"):
                        _merge_user_into(bible_users, user, data["dates_old"], data["emoji"], leader)
                    if data.get("dates_new"):
                        _merge_user_into(nt_users, user, data["dates_new"], data["emoji"], leader)
        elif schedule_type == "education":
            classified = _classify_education_users(users, edu_config)
            for user, data in classified["bible"].items():
                _merge_user_into(bible_users, user, data["dates"], data["emoji"], leader)
            for user, data in classified["nt"].items():
                _merge_user_into(nt_users, user, data["dates"], data["emoji"], leader)
        else:
            # unknown → 성경일독 기본값
            for user, data in users.items():
                _merge_user_into(bible_users, user, data["dates"], data["emoji"], leader)

        file_date = _extract_date_from_filename(file_name)
        if file_date and (oldest_file_date is None or file_date < oldest_file_date):
            oldest_file_date = file_date

        processed_rooms.append(room_name or file_name)

    logger.info("통합 완료: 성경일독 %d명, 신약일독 %d명, 투트랙 %d명, %d개 방, %d개 스킵",
                len(bible_users), len(nt_users), len(dual_users),
                len(processed_rooms), len(skipped_files))

    return {
        "success": True,
        "bible_users": bible_users,
        "nt_users": nt_users,
        "dual_users": dual_users,
        "processed_rooms": processed_rooms,
        "skipped_files": skipped_files,
        "oldest_file_date": oldest_file_date,
    }


def build_merged_xlsx(bible_users, nt_users, dual_users=None):
    """통합 XLSX 파일을 생성한다.

    Args:
        bible_users: {user: {"dates": set, "emoji": str, "leader": str}}
        nt_users: {user: {"dates": set, "emoji": str, "leader": str}}
        dual_users: {user: {"dates_old": set, "dates_new": set, "emoji": str, "leader": str}} (선택)

    Returns:
        bytes: XLSX 파일 바이트
    """
    wb = Workbook()

    # 성경일독 시트
    ws_bible = wb.active
    ws_bible.title = "성경일독 진도표"
    _build_merged_sheet(ws_bible, bible_users, title="2026 꿀성경 통합 진도표")

    # 신약일독 시트
    ws_nt = wb.create_sheet(title="신약일독 진도표")
    _build_merged_sheet(ws_nt, nt_users, title="2026 꿀성경 통합 진도표")

    # 투트랙 시트 (구약/신약 분리)
    if dual_users:
        ws_dual = wb.create_sheet(title="투트랙 진도표")
        _build_merged_dual_sheet(ws_dual, dual_users, title="2026 꿀성경 통합 진도표")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _apply_leader_merge(ws, rows, title):
    """같은 담당자의 셀을 병합하고 그룹 경계선을 복원한다."""
    if not rows:
        return
    data_start = (3 if title else 2) + ROW_PAD
    leader_col_ws = 1 + COL_PAD

    merge_start = data_start
    prev_leader = rows[0][0]
    for i in range(1, len(rows)):
        current_leader = rows[i][0]
        if current_leader != prev_leader:
            merge_end = data_start + i - 1
            if merge_end > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=leader_col_ws,
                               end_row=merge_end, end_column=leader_col_ws)
            merge_start = data_start + i
            prev_leader = current_leader
    merge_end = data_start + len(rows) - 1
    if merge_end > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=leader_col_ws,
                       end_row=merge_end, end_column=leader_col_ws)

    # 병합 후 담당 열 그룹 경계선 복원
    medium_side = Side(style="medium")
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col == leader_col_ws:
            cell = ws.cell(row=rng.min_row, column=leader_col_ws)
            cell.border = Border(
                top=cell.border.top,
                bottom=medium_side,
                left=cell.border.left,
                right=cell.border.right,
            )
            if rng.max_row > rng.min_row:
                bottom_cell = ws.cell(row=rng.max_row, column=leader_col_ws)
                bottom_cell.border = Border(
                    bottom=medium_side,
                    left=bottom_cell.border.left,
                    right=bottom_cell.border.right,
                )


def _build_merged_sheet(ws, users, title=None):
    """통합 시트 하나를 생성한다."""
    all_dates = set()
    for data in users.values():
        all_dates.update(data["dates"])
    all_dates_sorted = sort_dates(all_dates)

    headers = ["담당", "이름", "이모티콘"] + all_dates_sorted

    # 사용자 정렬: 담당 → 이름 순
    sorted_users = sorted(users.keys(), key=lambda u: (users[u].get("leader", ""), u))

    rows = []
    for user in sorted_users:
        data = users[user]
        if not data["dates"]:
            continue
        row = [data.get("leader", ""), user, data["emoji"]]
        row.extend("O" if d in data["dates"] else "" for d in all_dates_sorted)
        rows.append(row)

    _apply_sheet_style(ws, headers, rows, leader_col=1, title=title)
    _apply_leader_merge(ws, rows, title)
    if title:
        stats_text = _compute_stats(users, all_dates_sorted)
        _insert_stats_row(ws, stats_text, len(headers))


def _build_merged_dual_sheet(ws, dual_users, title=None):
    """투트랙 통합 시트를 생성한다 (사용자별 구약/신약 행 분리)."""
    all_dates = set()
    for data in dual_users.values():
        all_dates.update(data["dates_old"])
        all_dates.update(data["dates_new"])
    all_dates_sorted = sort_dates(all_dates)

    headers = ["담당", "이름", "이모티콘", "트랙"] + all_dates_sorted

    sorted_users = sorted(dual_users.keys(), key=lambda u: (dual_users[u].get("leader", ""), u))

    rows = []
    for user in sorted_users:
        data = dual_users[user]
        if data["dates_old"]:
            row = [data.get("leader", ""), user, data["emoji"], "구약"]
            row.extend("O" if d in data["dates_old"] else "" for d in all_dates_sorted)
            rows.append(row)
        if data["dates_new"]:
            row = [data.get("leader", ""), user, data["emoji"], "신약"]
            row.extend("O" if d in data["dates_new"] else "" for d in all_dates_sorted)
            rows.append(row)

    _apply_sheet_style(ws, headers, rows, leader_col=1, title=title)
    _apply_leader_merge(ws, rows, title)
    if title:
        stats_text = _compute_dual_stats(dual_users, all_dates_sorted)
        _insert_stats_row(ws, stats_text, len(headers))


def build_merged_preview(bible_users, nt_users, dual_users=None):
    """통합 미리보기 데이터를 생성한다.

    Returns:
        tuple: (headers, rows) — 성경일독 + 신약일독 + 투트랙 합쳐서 미리보기용
    """
    all_dates = set()
    for data in bible_users.values():
        all_dates.update(data["dates"])
    for data in nt_users.values():
        all_dates.update(data["dates"])
    if dual_users:
        for data in dual_users.values():
            all_dates.update(data.get("dates_old", set()))
            all_dates.update(data.get("dates_new", set()))
    all_dates_sorted = sort_dates(all_dates)

    headers = ["담당", "이름", "이모티콘", "트랙"] + all_dates_sorted

    rows = []
    # 성경일독
    for user in sorted(bible_users.keys(), key=lambda u: (bible_users[u].get("leader", ""), u)):
        data = bible_users[user]
        if not data["dates"]:
            continue
        row = [data.get("leader", ""), user, data["emoji"], "성경일독"]
        row.extend("O" if d in data["dates"] else "" for d in all_dates_sorted)
        rows.append(row)

    # 신약일독
    for user in sorted(nt_users.keys(), key=lambda u: (nt_users[u].get("leader", ""), u)):
        data = nt_users[user]
        if not data["dates"]:
            continue
        row = [data.get("leader", ""), user, data["emoji"], "신약일독"]
        row.extend("O" if d in data["dates"] else "" for d in all_dates_sorted)
        rows.append(row)

    # 투트랙 (구약/신약 분리)
    if dual_users:
        for user in sorted(dual_users.keys(), key=lambda u: (dual_users[u].get("leader", ""), u)):
            data = dual_users[user]
            if data.get("dates_old"):
                row = [data.get("leader", ""), user, data["emoji"], "투트랙(구약)"]
                row.extend("O" if d in data["dates_old"] else "" for d in all_dates_sorted)
                rows.append(row)
            if data.get("dates_new"):
                row = [data.get("leader", ""), user, data["emoji"], "투트랙(신약)"]
                row.extend("O" if d in data["dates_new"] else "" for d in all_dates_sorted)
                rows.append(row)

    return headers, rows
