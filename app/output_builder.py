"""CSV/XLSX/미리보기 출력 생성 모듈. analyzer.py에서 분리."""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.completion import (
    TRACK_LABELS,
    completion_row,
    expected_dates,
    is_complete,
    normalize_part,
    single_track_for_user,
)
from app.logger import get_logger
from app.style_constants import COL_PAD, ROW_PAD

logger = get_logger("output_builder")


def sort_dates(dates):
    def key(value):
        try:
            month, day = value.split("/")
            return (int(month), int(day))
        except (ValueError, TypeError):
            return (99, 99)

    return sorted(dates, key=key)


def build_output_csv(users, track_mode="single"):
    output = io.StringIO(newline="")
    writer = csv.writer(output)

    if track_mode == "dual":
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry.get("dates_old", set()))
            all_dates.update(entry.get("dates_new", set()))
        all_dates_sorted = sort_dates(all_dates)

        header = ["이름", "이모티콘", "트랙"]
        header.extend(all_dates_sorted)
        writer.writerow(header)

        all_users = sorted(users.keys())
        for user in all_users:
            entry = users[user]
            if entry.get("dates_old"):
                row = [user, entry.get("emoji", ""), "구약"]
                row.extend("O" if d in entry["dates_old"] else "" for d in all_dates_sorted)
                writer.writerow(row)
            if entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), "신약"]
                row.extend("O" if d in entry["dates_new"] else "" for d in all_dates_sorted)
                writer.writerow(row)
            if not entry.get("dates_old") and not entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), ""]
                row.extend("" for _ in all_dates_sorted)
                writer.writerow(row)
    else:
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry["dates"])

        all_dates_sorted = sort_dates(all_dates)

        header = ["이름", "이모티콘"]
        header.extend(all_dates_sorted)
        writer.writerow(header)

        for user in sorted(users.keys()):
            entry = users[user]
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates"] else "" for d in all_dates_sorted)
            writer.writerow(row)

    return output.getvalue().encode("utf-8-sig")


def build_preview_data(users, track_mode="single"):
    if track_mode == "dual":
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry.get("dates_old", set()))
            all_dates.update(entry.get("dates_new", set()))
        all_dates_sorted = sort_dates(all_dates)

        headers = ["이름", "이모티콘", "트랙"]
        headers.extend(all_dates_sorted)

        rows = []
        all_users = sorted(users.keys())
        for user in all_users:
            entry = users[user]
            if entry.get("dates_old"):
                row = [user, entry.get("emoji", ""), "구약"]
                row.extend("O" if d in entry["dates_old"] else "" for d in all_dates_sorted)
                rows.append(row)
            if entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), "신약"]
                row.extend("O" if d in entry["dates_new"] else "" for d in all_dates_sorted)
                rows.append(row)
            if not entry.get("dates_old") and not entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), ""]
                row.extend("" for _ in all_dates_sorted)
                rows.append(row)
    else:
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry["dates"])

        all_dates_sorted = sort_dates(all_dates)

        headers = ["이름", "이모티콘"]
        headers.extend(all_dates_sorted)

        rows = []
        for user in sorted(users.keys()):
            entry = users[user]
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates"] else "" for d in all_dates_sorted)
            rows.append(row)

    return headers, rows


def apply_sheet_style(ws, headers, rows, leader_col=None, title=None, completed_rows=None):
    """XLSX 시트에 스타일(헤더, 데이터, 테두리, 고정 틀 등)을 적용한다."""
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11)
    body_font = Font(name="맑은 고딕", size=11)
    mark_font = Font(name="맑은 고딕", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    name_fill = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
    completed_name_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    completed_row_offsets = set(completed_rows or [])

    thin = Side(style="thin")
    medium = Side(style="medium")

    R = ROW_PAD
    C = COL_PAD
    first_col = 1 + C
    last_col = len(headers) + C

    name_col = (headers.index("이름") + 1 + C) if "이름" in headers else first_col
    _NON_DATE = {"이름", "이모티콘", "담당", "트랙", "완독일수", "전체일수"}
    date_start_col = sum(1 for h in headers if h in _NON_DATE) + 1 + C
    leader_col_ws = (leader_col + C) if leader_col else None

    # 패딩 행/열 크기
    ws.column_dimensions['A'].width = 2
    ws.row_dimensions[1].height = 6

    # 타이틀 행 오프셋
    if title:
        title_row = 1 + R
        header_row = 2 + R
        data_start = 3 + R
        freeze = "B4"
    else:
        header_row = 1 + R
        data_start = 2 + R
        freeze = "B3"

    last_row = data_start + len(rows) - 1 if rows else header_row

    # 담당자 그룹 경계 행 (leader_col이 있을 때만)
    leader_boundary_rows = set()
    if leader_col and rows:
        for i in range(len(rows) - 1):
            lc = rows[i][leader_col - 1] if (leader_col - 1) < len(rows[i]) else ""
            ln = rows[i + 1][leader_col - 1] if (leader_col - 1) < len(rows[i + 1]) else ""
            if lc != ln:
                leader_boundary_rows.add(data_start + i)

    # 타이틀 행
    if title:
        ws.merge_cells(start_row=title_row, start_column=first_col,
                        end_row=title_row, end_column=last_col)
        title_cell = ws.cell(row=title_row, column=first_col, value=title)
        title_cell.font = Font(name="맑은 고딕", size=20)
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        ws.row_dimensions[title_row].height = 60
        # 타이틀 행 외곽 medium
        title_cell.border = Border(top=medium, bottom=medium, left=medium, right=medium)
        for c in range(first_col + 1, last_col + 1):
            cell = ws.cell(row=title_row, column=c)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = Border(
                top=medium,
                bottom=medium,
                left=thin,
                right=medium if c == last_col else thin,
            )

    # 헤더 행
    ws.row_dimensions[header_row].height = 28
    for col_idx, value in enumerate(headers, start=first_col):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 데이터 행
    for row_offset, row_data in enumerate(rows):
        row_idx = data_start + row_offset
        for col_idx, value in enumerate(row_data, start=first_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = mark_font if value == "O" else body_font
            cell.alignment = center_align
            if col_idx == name_col:
                cell.fill = completed_name_fill if row_offset in completed_row_offsets else name_fill
            elif leader_col_ws and col_idx == leader_col_ws:
                cell.fill = name_fill

    # 풀 그리드 테두리
    for r in range(header_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            top_side = medium if r == header_row else thin
            bottom_side = (
                medium if r in (header_row, last_row) or r in leader_boundary_rows
                else thin
            )
            left_side = medium if c in (first_col, date_start_col) else thin
            right_side = medium if c == last_col else thin
            cell.border = Border(top=top_side, bottom=bottom_side,
                                 left=left_side, right=right_side)

    # 열 너비
    for col_idx, header in enumerate(headers, start=first_col):
        col_letter = ws.cell(row=header_row, column=col_idx).column_letter
        if header == "이름":
            ws.column_dimensions[col_letter].width = 16
        elif header == "이모티콘":
            ws.column_dimensions[col_letter].width = 8
        elif header == "담당":
            ws.column_dimensions[col_letter].width = 10
        elif header == "트랙":
            ws.column_dimensions[col_letter].width = 14
        elif header in ("완독일수", "전체일수"):
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 7

    ws.freeze_panes = freeze


def build_dual_preview_data(users):
    """Dual 모드에서 구약/신약 별도의 headers+rows를 반환한다."""
    all_users = sorted(users.keys())

    old_dates = set()
    new_dates = set()
    for entry in users.values():
        old_dates.update(entry.get("dates_old", set()))
        new_dates.update(entry.get("dates_new", set()))
    old_dates_sorted = sort_dates(old_dates)
    new_dates_sorted = sort_dates(new_dates)

    old_headers = ["이름", "이모티콘"] + old_dates_sorted
    old_rows = []
    for user in all_users:
        entry = users[user]
        if entry.get("dates_old"):
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates_old"] else "" for d in old_dates_sorted)
            old_rows.append(row)

    new_headers = ["이름", "이모티콘"] + new_dates_sorted
    new_rows = []
    for user in all_users:
        entry = users[user]
        if entry.get("dates_new"):
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates_new"] else "" for d in new_dates_sorted)
            new_rows.append(row)

    return old_headers, old_rows, new_headers, new_rows


def _add_meta_sheet(wb, meta):
    """숨겨진 _메타 시트를 생성하여 key-value 메타데이터를 저장한다."""
    ws = wb.create_sheet(title="_메타")
    ws.sheet_state = "hidden"
    for row_idx, (key, value) in enumerate(meta.items(), start=1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=str(value))


def _completion_part(meta):
    return normalize_part((meta or {}).get("part", 1))


def _build_completion_sheet(wb, rows):
    """완독자 리스트 시트를 추가한다."""
    headers = ["트랙", "이름", "이모티콘", "완독일수", "전체일수"]
    ws = wb.create_sheet(title="완독자")
    apply_sheet_style(ws, headers, rows, completed_rows=range(len(rows)))


def _single_completion_data(users, rows, meta):
    """싱글 분석 XLSX용 완독자 행과 강조 대상 행을 계산한다."""
    schedule_type = (meta or {}).get("schedule_type", "bible")
    part = _completion_part(meta)
    completed_rows = []
    completion_rows = []

    for row_idx, row in enumerate(rows):
        if not row:
            continue
        user = row[0]
        data = users.get(user, {})
        track = single_track_for_user(user, schedule_type)
        if not track:
            continue
        expected = expected_dates(track, part)
        dates = data.get("dates", set())
        if is_complete(dates, expected):
            completed_rows.append(row_idx)
        row_data = completion_row(
            TRACK_LABELS.get(track, track),
            user,
            data.get("emoji", ""),
            dates,
            expected,
        )
        if row_data:
            completion_rows.append(row_data)

    return completed_rows, completion_rows


def _dual_completion_data(users, old_rows, new_rows, meta):
    """투트랙 분석 XLSX용 완독자 행과 강조 대상 행을 계산한다."""
    part = _completion_part(meta)
    old_expected = expected_dates("old", part)
    new_expected = expected_dates("new", part)

    old_completed_rows = []
    new_completed_rows = []
    completion_rows = []

    for row_idx, row in enumerate(old_rows):
        user = row[0]
        if is_complete(users.get(user, {}).get("dates_old", set()), old_expected):
            old_completed_rows.append(row_idx)

    for row_idx, row in enumerate(new_rows):
        user = row[0]
        if is_complete(users.get(user, {}).get("dates_new", set()), new_expected):
            new_completed_rows.append(row_idx)

    for user in sorted(users.keys()):
        data = users[user]
        dates_old = data.get("dates_old", set())
        dates_new = data.get("dates_new", set())
        old_complete = is_complete(dates_old, old_expected)
        new_complete = is_complete(dates_new, new_expected)

        old_row = completion_row("구약", user, data.get("emoji", ""), dates_old, old_expected)
        if old_row:
            completion_rows.append(old_row)
        new_row = completion_row("신약", user, data.get("emoji", ""), dates_new, new_expected)
        if new_row:
            completion_rows.append(new_row)
        if old_complete and new_complete:
            total_count = len(old_expected) + len(new_expected)
            completion_rows.append(["둘 다", user, data.get("emoji", ""), total_count, total_count])

    return old_completed_rows, new_completed_rows, completion_rows


def build_output_xlsx(users, track_mode="single", meta=None):
    wb = Workbook()

    if track_mode == "dual":
        old_headers, old_rows, new_headers, new_rows = build_dual_preview_data(users)
        old_completed_rows, new_completed_rows, completion_rows = _dual_completion_data(
            users, old_rows, new_rows, meta
        )

        ws_old = wb.active
        ws_old.title = "구약 진도표"
        apply_sheet_style(ws_old, old_headers, old_rows, completed_rows=old_completed_rows)

        ws_new = wb.create_sheet(title="신약 진도표")
        apply_sheet_style(ws_new, new_headers, new_rows, completed_rows=new_completed_rows)
        _build_completion_sheet(wb, completion_rows)
    else:
        headers, rows = build_preview_data(users, track_mode)
        completed_rows, completion_rows = _single_completion_data(users, rows, meta)
        ws = wb.active
        ws.title = "꿀성경 진도표"
        apply_sheet_style(ws, headers, rows, completed_rows=completed_rows)
        _build_completion_sheet(wb, completion_rows)

    if meta:
        _add_meta_sheet(wb, meta)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
