import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.date_parser import DATE_TIME_PATTERN, parse_dates
from app.emoji import extract_trailing_emoji, is_emoji_component, normalize_emoji
from app.logger import get_logger
from app.schedule import BIBLE_DATES, NT_DATES, detect_schedule

logger = get_logger("analyzer")

MAX_DATES_PER_MESSAGE = 14

_STRIP_WORDS = ("광천", "누나", "오빠", "언니")


def normalize_user_name(name: str) -> str:
    """이름에서 숫자·이모티콘·영어·공백·특정 키워드를 제거하여 한글 이름만 남긴다."""
    result = name
    for word in _STRIP_WORDS:
        result = result.replace(word, "")
    result = "".join(
        ch for ch in result
        if not is_emoji_component(ch)
        and not ch.isdigit()
        and not ch.isspace()
        and not ("A" <= ch <= "Z" or "a" <= ch <= "z")
    )
    return result if result else name


def _max_date(dates):
    best = None
    for d in dates:
        try:
            month, day = d.split("/")
            t = (int(month), int(day))
        except (ValueError, TypeError):
            continue
        if best is None or t > best:
            best = t
    return best


def decode_payload(payload):
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            result = payload.decode(encoding)
            logger.info("인코딩 감지: %s (%d bytes → %d chars)", encoding, len(payload), len(result))
            return result
        except UnicodeDecodeError:
            continue
    logger.warning("인코딩 감지 실패 — utf-8 errors=replace로 폴백 (%d bytes)", len(payload))
    return payload.decode("utf-8", errors="replace")


# 하위호환 alias
decode_csv_payload = decode_payload


def iter_data_rows(reader):
    first_row = next(reader, None)
    if first_row is None:
        return
    if first_row and not DATE_TIME_PATTERN.match(first_row[0].strip()):
        pass
    else:
        yield first_row
    for row in reader:
        yield row


def choose_assigned_emoji(counts, order):
    if not counts:
        return ""
    max_count = max(counts.values())
    for emoji in order:
        if counts.get(emoji) == max_count:
            return emoji
    return next(iter(counts.keys()))


def message_contains_emoji(message, emoji_key, emoji_raw):
    if emoji_raw and emoji_raw in message:
        return True
    normalized_message = normalize_emoji(message)
    return emoji_key in normalized_message


def extract_tracks(message):
    tracks = set()
    if "구약" in message:
        tracks.add("old")
    if "신약" in message:
        tracks.add("new")
    return tracks


def parse_csv_rows(csv_text):
    """CSV 텍스트를 파싱하여 (user, message) 튜플 리스트를 반환한다."""
    rows = []
    total_rows = 0
    skip_short = 0
    skip_empty = 0
    reader = csv.reader(io.StringIO(csv_text, newline=""))
    for row in iter_data_rows(reader):
        total_rows += 1
        if not row or len(row) < 3:
            skip_short += 1
            continue
        user = row[1].strip()
        message = row[2].strip()
        if not user or not message:
            skip_empty += 1
            continue
        rows.append((user, message))
    logger.info("CSV 파싱: 전체 %d행, 유효 %d행 (컬럼 부족 %d행, 빈 값 %d행 스킵)",
                total_rows, len(rows), skip_short, skip_empty)
    return rows


def analyze_chat(csv_text=None, track_mode="single", rows=None):
    if rows is None:
        rows = parse_csv_rows(csv_text or "")

    # 사용자 이름 정규화
    rows = [(normalize_user_name(user), message) for user, message in rows]

    logger.info("파싱된 메시지 수: %d, 트랙 모드: %s", len(rows), track_mode)
    if not rows:
        logger.warning("파싱된 메시지가 0건 — 파일 형식이나 인코딩을 확인하세요")
        return {}

    unique_users = {user for user, _ in rows}
    logger.info("메시지 발신자 수: %d명 (%s)", len(unique_users),
                ", ".join(sorted(unique_users)[:10]) + ("..." if len(unique_users) > 10 else ""))

    emoji_counts = {}
    emoji_order = {}
    emoji_raw = {}

    skip_no_date = 0
    skip_too_many_dates = 0
    skip_no_emoji = 0

    for user, message in rows:
        dates = parse_dates(message)
        if not dates:
            skip_no_date += 1
            continue
        if len(dates) > MAX_DATES_PER_MESSAGE:
            skip_too_many_dates += 1
            continue
        trailing_emoji = extract_trailing_emoji(message)
        if not trailing_emoji:
            skip_no_emoji += 1
            continue
        emoji_key = normalize_emoji(trailing_emoji)
        counts = emoji_counts.setdefault(user, {})
        counts[emoji_key] = counts.get(emoji_key, 0) + 1
        order = emoji_order.setdefault(user, [])
        if emoji_key not in order:
            order.append(emoji_key)
        raw_map = emoji_raw.setdefault(user, {})
        if emoji_key not in raw_map:
            raw_map[emoji_key] = trailing_emoji

    logger.info(
        "이모지 감지 단계 — 날짜 없음: %d건, 날짜 과다(>%d): %d건, 이모지 없음: %d건",
        skip_no_date, MAX_DATES_PER_MESSAGE, skip_too_many_dates, skip_no_emoji,
    )

    user_emojis = {}
    for user, counts in emoji_counts.items():
        order = emoji_order.get(user, [])
        emoji_key = choose_assigned_emoji(counts, order)
        if not emoji_key:
            continue
        emoji_value = emoji_raw.get(user, {}).get(emoji_key, emoji_key)
        user_emojis[user] = {"emoji_key": emoji_key, "emoji": emoji_value}

    logger.info("이모지 할당된 사용자: %d명", len(user_emojis))
    if not user_emojis:
        logger.warning(
            "이모지 할당된 사용자가 0명 — 메시지에 날짜+이모지 조합이 없습니다. "
            "샘플 메시지(처음 5건): %s",
            [msg[:80] for _, msg in rows[:5]],
        )
        return {}
    for user, info in sorted(user_emojis.items()):
        logger.debug("  %s → %s", user, info["emoji"])

    if track_mode == "dual":
        schedule_old = BIBLE_DATES
        schedule_new = NT_DATES
        logger.info("듀얼 모드 — 구약 진도표 날짜 %d개, 신약 진도표 날짜 %d개",
                     len(schedule_old), len(schedule_new))
    else:
        schedule = detect_schedule(rows)
        if schedule is not None:
            logger.info("싱글 모드 — 진도표 감지됨 (유효 날짜 %d개)", len(schedule))
        else:
            logger.info("싱글 모드 — 진도표 미감지 (날짜 필터 미적용)")

    users = {}
    user_last_date = {}
    skip_no_assigned = 0
    skip_emoji_mismatch = 0
    skip_no_dates_2 = 0
    skip_too_many_dates_2 = 0
    skip_schedule_filter = 0
    skip_no_track = 0

    for user, message in rows:
        assigned = user_emojis.get(user)
        if not assigned:
            skip_no_assigned += 1
            continue
        if not message_contains_emoji(message, assigned["emoji_key"], assigned["emoji"]):
            skip_emoji_mismatch += 1
            continue

        if track_mode == "dual":
            tracks = extract_tracks(message)
            if not tracks:
                skip_no_track += 1
                continue
            last_old = user_last_date.get((user, "old"))
            last_new = user_last_date.get((user, "new"))
            if tracks == {"old"}:
                last_date = last_old
            elif tracks == {"new"}:
                last_date = last_new
            else:
                if last_old is not None and last_new is not None:
                    last_date = min(last_old, last_new)
                else:
                    last_date = last_old or last_new
            dates = parse_dates(message, last_date=last_date)
        else:
            last_date = user_last_date.get(user)
            dates = parse_dates(message, last_date=last_date)

        if not dates:
            skip_no_dates_2 += 1
            continue
        if len(dates) > MAX_DATES_PER_MESSAGE:
            skip_too_many_dates_2 += 1
            continue

        if track_mode == "dual":
            dates_old = [d for d in dates if d in schedule_old] if "old" in tracks else []
            dates_new = [d for d in dates if d in schedule_new] if "new" in tracks else []
            if not dates_old and not dates_new:
                skip_schedule_filter += 1
                continue
            entry = users.setdefault(
                user,
                {"dates_old": set(), "dates_new": set(), "emoji": assigned["emoji"]},
            )
            for date_value in dates_old:
                entry["dates_old"].add(date_value)
            for date_value in dates_new:
                entry["dates_new"].add(date_value)
            if dates_old:
                md = _max_date(dates_old)
                if md:
                    prev = user_last_date.get((user, "old"))
                    if prev is None or md > prev:
                        user_last_date[(user, "old")] = md
            if dates_new:
                md = _max_date(dates_new)
                if md:
                    prev = user_last_date.get((user, "new"))
                    if prev is None or md > prev:
                        user_last_date[(user, "new")] = md
        else:
            if schedule is not None:
                filtered_dates = [d for d in dates if d in schedule]
                if not filtered_dates and dates:
                    skip_schedule_filter += 1
                dates = filtered_dates
            if not dates:
                continue
            entry = users.setdefault(
                user,
                {"dates": set(), "emoji": assigned["emoji"]},
            )
            for date_value in dates:
                entry["dates"].add(date_value)
            md = _max_date(dates)
            if md:
                prev = user_last_date.get(user)
                if prev is None or md > prev:
                    user_last_date[user] = md

    logger.info(
        "날짜 수집 단계 — 이모지 미할당: %d건, 이모지 불일치: %d건, "
        "날짜 없음: %d건, 날짜 과다: %d건, 진도표 필터링: %d건%s",
        skip_no_assigned, skip_emoji_mismatch,
        skip_no_dates_2, skip_too_many_dates_2, skip_schedule_filter,
        f", 트랙 미지정(구약/신약): {skip_no_track}건" if track_mode == "dual" else "",
    )
    logger.info("최종 분석 결과: %d명", len(users))
    if not users:
        logger.warning(
            "분석 결과 0명 — 이모지 할당자 %d명 중 유효 날짜를 가진 사용자가 없습니다",
            len(user_emojis),
        )

    return users


def sort_dates(dates):
    def key(value):
        try:
            month, day = value.split("/")
            return (int(month), int(day))
        except (ValueError, TypeError):
            return (99, 99)

    return sorted(dates, key=key)


def build_output_csv(users, track_mode="single"):
    output = io.StringIO(newline="")
    writer = csv.writer(output)

    if track_mode == "dual":
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry.get("dates_old", set()))
            all_dates.update(entry.get("dates_new", set()))
        all_dates_sorted = sort_dates(all_dates)

        header = ["이름", "이모티콘", "트랙"]
        header.extend(all_dates_sorted)
        writer.writerow(header)

        all_users = sorted(
            u for u, e in users.items()
            if e.get("dates_old") or e.get("dates_new")
        )
        for user in all_users:
            entry = users[user]
            if entry.get("dates_old"):
                row = [user, entry.get("emoji", ""), "구약"]
                row.extend("O" if d in entry["dates_old"] else "" for d in all_dates_sorted)
                writer.writerow(row)
            if entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), "신약"]
                row.extend("O" if d in entry["dates_new"] else "" for d in all_dates_sorted)
                writer.writerow(row)
    else:
        user_dates = {}
        all_dates = set()
        for user, entry in users.items():
            if not entry["dates"]:
                continue
            user_dates[user] = entry["dates"]
            all_dates.update(entry["dates"])

        all_dates_sorted = sort_dates(all_dates)

        header = ["이름", "이모티콘"]
        header.extend(all_dates_sorted)
        writer.writerow(header)

        for user in sorted(user_dates.keys()):
            entry = users[user]
            date_set = user_dates[user]
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in date_set else "" for d in all_dates_sorted)
            writer.writerow(row)

    return output.getvalue().encode("utf-8-sig")


def build_preview_data(users, track_mode="single"):
    if track_mode == "dual":
        all_dates = set()
        for entry in users.values():
            all_dates.update(entry.get("dates_old", set()))
            all_dates.update(entry.get("dates_new", set()))
        all_dates_sorted = sort_dates(all_dates)

        headers = ["이름", "이모티콘", "트랙"]
        headers.extend(all_dates_sorted)

        rows = []
        all_users = sorted(
            u for u, e in users.items()
            if e.get("dates_old") or e.get("dates_new")
        )
        for user in all_users:
            entry = users[user]
            if entry.get("dates_old"):
                row = [user, entry.get("emoji", ""), "구약"]
                row.extend("O" if d in entry["dates_old"] else "" for d in all_dates_sorted)
                rows.append(row)
            if entry.get("dates_new"):
                row = [user, entry.get("emoji", ""), "신약"]
                row.extend("O" if d in entry["dates_new"] else "" for d in all_dates_sorted)
                rows.append(row)
    else:
        user_dates = {}
        all_dates = set()
        for user, entry in users.items():
            if not entry["dates"]:
                continue
            user_dates[user] = entry["dates"]
            all_dates.update(entry["dates"])

        all_dates_sorted = sort_dates(all_dates)

        headers = ["이름", "이모티콘"]
        headers.extend(all_dates_sorted)

        rows = []
        for user in sorted(user_dates.keys()):
            entry = users[user]
            date_set = user_dates[user]
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in date_set else "" for d in all_dates_sorted)
            rows.append(row)

    return headers, rows


def _apply_sheet_style(ws, headers, rows):
    header_fill = PatternFill(start_color="FFF6E2", end_color="FFF6E2", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="4A2D14")
    header_border = Border(bottom=Side(style="medium", color="C46B12"))

    body_font = Font(name="맑은 고딕", size=11, color="2A1A08")
    mark_font = Font(name="맑은 고딕", bold=True, size=11, color="E39B2F")
    row_border = Border(bottom=Side(style="thin", color="D4C4A8"))
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = center_align

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if value == "O":
                cell.font = mark_font
            else:
                cell.font = body_font
            cell.border = row_border
            cell.alignment = center_align

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 8
    date_start_col = 3
    for col_idx in range(date_start_col, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 7

    ws.freeze_panes = "A2"


def build_dual_preview_data(users):
    """Dual 모드에서 구약/신약 별도의 headers+rows를 반환한다."""
    all_users = sorted(
        u for u, e in users.items()
        if e.get("dates_old") or e.get("dates_new")
    )

    old_dates = set()
    new_dates = set()
    for entry in users.values():
        old_dates.update(entry.get("dates_old", set()))
        new_dates.update(entry.get("dates_new", set()))
    old_dates_sorted = sort_dates(old_dates)
    new_dates_sorted = sort_dates(new_dates)

    old_headers = ["이름", "이모티콘"] + old_dates_sorted
    old_rows = []
    for user in all_users:
        entry = users[user]
        if entry.get("dates_old"):
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates_old"] else "" for d in old_dates_sorted)
            old_rows.append(row)

    new_headers = ["이름", "이모티콘"] + new_dates_sorted
    new_rows = []
    for user in all_users:
        entry = users[user]
        if entry.get("dates_new"):
            row = [user, entry.get("emoji", "")]
            row.extend("O" if d in entry["dates_new"] else "" for d in new_dates_sorted)
            new_rows.append(row)

    return old_headers, old_rows, new_headers, new_rows


def _add_meta_sheet(wb, meta):
    """숨겨진 _메타 시트를 생성하여 key-value 메타데이터를 저장한다."""
    ws = wb.create_sheet(title="_메타")
    ws.sheet_state = "hidden"
    for row_idx, (key, value) in enumerate(meta.items(), start=1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=str(value))


def build_output_xlsx(users, track_mode="single", meta=None):
    wb = Workbook()

    if track_mode == "dual":
        old_headers, old_rows, new_headers, new_rows = build_dual_preview_data(users)

        ws_old = wb.active
        ws_old.title = "구약 진도표"
        _apply_sheet_style(ws_old, old_headers, old_rows)

        ws_new = wb.create_sheet(title="신약 진도표")
        _apply_sheet_style(ws_new, new_headers, new_rows)
    else:
        headers, rows = build_preview_data(users, track_mode)
        ws = wb.active
        ws.title = "꿀성경 진도표"
        _apply_sheet_style(ws, headers, rows)

    if meta:
        _add_meta_sheet(wb, meta)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
