"""분석결과 시트 생성 및 진행 통계 계산."""

import datetime
from dataclasses import dataclass
from math import ceil
from statistics import median

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter, quote_sheetname

from app.completion import expected_dates, is_complete, normalize_part, single_track_for_user
from app.schedule import get_part_books
from app.style_constants import COL_PAD, ROW_PAD

GROUP_ALL = "전체"
GROUP_BIBLE = "성경일독"
GROUP_NT = "신약일독"
GROUP_DUAL = "투트랙"
GROUPS = [GROUP_ALL, GROUP_BIBLE, GROUP_NT, GROUP_DUAL]

PROGRESS_BUCKETS = ["0%", "1~25%", "26~50%", "51~75%", "76~99%", "100%"]
FORMULA_HELPER_SHEET = "_분석계산"
DROPOUT_VISIBLE_LIMIT = 8


@dataclass(frozen=True)
class AnalysisRecord:
    group: str
    name: str
    emoji: str
    leader: str
    read_count: int
    expected_count: int
    complete: bool
    last_date: str
    last_track: str
    last_position: str
    status: str
    activity_dates: frozenset

    @property
    def progress_rate(self):
        if self.expected_count == 0:
            return 0.0
        return min(1.0, self.read_count / self.expected_count)


@dataclass(frozen=True)
class AnalysisSourceRange:
    sheet_name: str
    header_row: int
    row: int
    date_start_col: int
    date_end_col: int
    track: str


def _date_key(value):
    try:
        month, day = str(value).split("/")
        return int(month), int(day)
    except (ValueError, TypeError):
        return 99, 99


def _date_value(value):
    try:
        month, day = str(value).split("/")
        return datetime.date(2026, int(month), int(day))
    except (ValueError, TypeError):
        return None


def _week_bucket(value):
    date_value = _date_value(value)
    if date_value is None:
        return datetime.date(9999, 12, 31), "미확인"
    week_start = date_value - datetime.timedelta(days=date_value.weekday())
    week_end = week_start + datetime.timedelta(days=6)
    return week_start, f"{week_start.month}/{week_start.day}~{week_end.month}/{week_end.day}"


def _sort_dates(dates):
    return sorted((d for d in dates if d), key=_date_key)


def _latest_date(dates):
    sorted_dates = _sort_dates(dates)
    return sorted_dates[-1] if sorted_dates else ""


def _count_valid_dates(dates, expected):
    return len(set(dates or ()) & set(expected or ()))


def _position_for_date(date_value, track, part):
    if not date_value:
        return "미시작"

    expected = _sort_dates(expected_dates(track, part))
    if not expected:
        return "진도 미확인"

    target_key = _date_key(date_value)
    index = None
    for i, expected_date in enumerate(expected):
        if _date_key(expected_date) <= target_key:
            index = i
        else:
            break
    if index is None:
        index = 0

    books = get_part_books("nt" if track == "nt" else "bible", part)
    if not books:
        return f"PART {normalize_part(part)} {index + 1}일차"

    ratio = index / max(1, len(expected) - 1)
    book_index = min(len(books) - 1, int(ratio * len(books)))
    return f"{books[book_index]} 부근"


def _status(read_count, complete):
    if complete:
        return "완독"
    if read_count == 0:
        return "미시작"
    return "하차 추정"


def _single_record(name, data, group, track, part, leader=""):
    expected = expected_dates(track, part)
    dates = set(data.get("dates", set()))
    read_count = _count_valid_dates(dates, expected)
    complete = is_complete(dates, expected)
    last_date = _latest_date(dates)
    return AnalysisRecord(
        group=group,
        name=name,
        emoji=data.get("emoji", ""),
        leader=leader or data.get("leader", ""),
        read_count=read_count,
        expected_count=len(expected),
        complete=complete,
        last_date=last_date,
        last_track=group,
        last_position=_position_for_date(last_date, track, part),
        status=_status(read_count, complete),
        activity_dates=frozenset(dates),
    )


def _dual_record(name, data, part, leader=""):
    old_expected = expected_dates("old", part)
    new_expected = expected_dates("new", part)
    old_dates = set(data.get("dates_old", set()))
    new_dates = set(data.get("dates_new", set()))
    old_read_count = _count_valid_dates(old_dates, old_expected)
    new_read_count = _count_valid_dates(new_dates, new_expected)
    old_complete = is_complete(old_dates, old_expected)
    new_complete = is_complete(new_dates, new_expected)

    old_last = _latest_date(old_dates)
    new_last = _latest_date(new_dates)
    old_key = _date_key(old_last) if old_last else None
    new_key = _date_key(new_last) if new_last else None

    if old_last and new_last and old_key == new_key:
        last_date = old_last
        last_track = "구약/신약"
        last_position = (
            f"구약: {_position_for_date(old_last, 'old', part)} / "
            f"신약: {_position_for_date(new_last, 'new', part)}"
        )
    elif new_last and (not old_last or new_key > old_key):
        last_date = new_last
        last_track = "신약"
        last_position = _position_for_date(new_last, "new", part)
    elif old_last:
        last_date = old_last
        last_track = "구약"
        last_position = _position_for_date(old_last, "old", part)
    else:
        last_date = ""
        last_track = ""
        last_position = "미시작"

    read_count = old_read_count + new_read_count
    complete = old_complete and new_complete
    return AnalysisRecord(
        group=GROUP_DUAL,
        name=name,
        emoji=data.get("emoji", ""),
        leader=leader or data.get("leader", ""),
        read_count=read_count,
        expected_count=len(old_expected) + len(new_expected),
        complete=complete,
        last_date=last_date,
        last_track=last_track,
        last_position=last_position,
        status=_status(read_count, complete),
        activity_dates=frozenset(old_dates | new_dates),
    )


def _dedupe_records(records, dedupe_names=None):
    """지정된 이름만 여러 그룹에 있을 때 대표 레코드 하나로 줄인다."""
    dedupe_names = set(dedupe_names or [])
    priority = {
        GROUP_DUAL: 3,
        GROUP_BIBLE: 2,
        GROUP_NT: 1,
    }

    def score(record):
        return (
            1 if record.complete else 0,
            record.progress_rate,
            record.read_count,
            priority.get(record.group, 0),
        )

    selected = {}
    kept = []
    for record in records:
        if record.name not in dedupe_names:
            kept.append(record)
            continue
        existing = selected.get(record.name)
        if existing is None or score(record) > score(existing):
            selected[record.name] = record
    kept.extend(selected.values())
    return kept


def build_output_analysis_records(users, track_mode="single", meta=None):
    """개별 분석 결과 XLSX용 사용자별 분석 레코드를 생성한다."""
    part = normalize_part((meta or {}).get("part", 1))
    if track_mode == "dual":
        return [
            _dual_record(name, users[name], part)
            for name in sorted(users.keys())
        ]

    schedule_type = (meta or {}).get("schedule_type", "bible")
    records = []
    for name in sorted(users.keys()):
        track = single_track_for_user(name, schedule_type)
        if not track:
            continue
        group = GROUP_NT if track == "nt" else GROUP_BIBLE
        records.append(_single_record(name, users[name], group, track, part))
    return records


def build_merged_analysis_records(bible_users, nt_users, dual_users=None, part=1, dedupe_names=None):
    """통합 XLSX용 사용자별 분석 레코드를 생성한다."""
    part = normalize_part(part)
    records = []

    for name in sorted(bible_users.keys(), key=lambda u: (bible_users[u].get("leader", ""), u)):
        records.append(_single_record(name, bible_users[name], GROUP_BIBLE, "bible", part))

    for name in sorted(nt_users.keys(), key=lambda u: (nt_users[u].get("leader", ""), u)):
        records.append(_single_record(name, nt_users[name], GROUP_NT, "nt", part))

    for name in sorted((dual_users or {}).keys(), key=lambda u: ((dual_users or {})[u].get("leader", ""), u)):
        records.append(_dual_record(name, dual_users[name], part))

    return sorted(_dedupe_records(records, dedupe_names), key=lambda r: (r.group, r.leader, r.name))


def dedupe_record_count(records, dedupe_names=None):
    """지정된 dedupe 대상 기준의 분석 레코드 인원 수를 반환한다."""
    return len(_dedupe_records(records, dedupe_names))


def dual_record_count(records):
    """투트랙 분석 레코드 수를 반환한다."""
    return sum(1 for record in records if record.group == GROUP_DUAL)


def summarize_records(records):
    """전체/그룹별 요약 행을 생성한다."""
    summary = []
    for group in GROUPS:
        selected = list(records) if group == GROUP_ALL else [r for r in records if r.group == group]
        total = len(selected)
        complete_count = sum(1 for r in selected if r.complete)
        not_started = sum(1 for r in selected if r.status == "미시작")
        dropout = sum(1 for r in selected if r.status == "하차 추정")
        rates = [r.progress_rate for r in selected]
        avg_rate = sum(rates) / total if total else 0.0
        median_rate = median(rates) if rates else 0.0
        summary.append({
            "group": group,
            "total": total,
            "complete": complete_count,
            "complete_rate": complete_count / total if total else 0.0,
            "avg_progress": avg_rate,
            "median_progress": median_rate,
            "not_started": not_started,
            "dropout": dropout,
        })
    return summary


def _bucket_for_rate(rate):
    if rate <= 0:
        return "0%"
    if rate >= 1:
        return "100%"
    percent = rate * 100
    if percent <= 25:
        return "1~25%"
    if percent <= 50:
        return "26~50%"
    if percent <= 75:
        return "51~75%"
    return "76~99%"


def progress_distribution(records):
    """진행률 구간별 인원 분포를 반환한다."""
    rows = []
    for bucket in PROGRESS_BUCKETS:
        row = {"bucket": bucket}
        for group in GROUPS:
            selected = records if group == GROUP_ALL else [r for r in records if r.group == group]
            row[group] = sum(1 for r in selected if _bucket_for_rate(r.progress_rate) == bucket)
        rows.append(row)
    return rows


def dropout_distribution(records, group=GROUP_ALL):
    """하차 추정자의 마지막 인증 주+진행 위치 분포를 반환한다."""
    selected = records if group == GROUP_ALL else [r for r in records if r.group == group]
    buckets = {}
    for record in selected:
        if record.status != "하차 추정":
            continue
        week_start, week_label = _week_bucket(record.last_date)
        key = (week_start, week_label, record.last_position)
        item = buckets.setdefault(key, {
            "week_start": week_start,
            "week": week_label,
            "position": record.last_position,
            "last_dates": set(),
            "names": [],
        })
        item["last_dates"].add(record.last_date)
        item["names"].append(record.name)

    rows = []
    for (_, week_label, position), item in sorted(buckets.items(), key=lambda entry: entry[0][0]):
        last_dates = _sort_dates(item["last_dates"])
        rows.append({
            "label": f"{week_label} | {position}",
            "week": week_label,
            "date": ", ".join(last_dates),
            "position": position,
            "count": len(item["names"]),
            "names": ", ".join(sorted(item["names"])),
        })
    return rows


def dropout_week_distribution(records, group=GROUP_ALL):
    """하차 추정자의 마지막 인증 주 단위 분포를 반환한다."""
    selected = records if group == GROUP_ALL else [r for r in records if r.group == group]
    buckets = {}
    for record in selected:
        if record.status != "하차 추정":
            continue
        week_start, week_label = _week_bucket(record.last_date)
        item = buckets.setdefault(week_start, {
            "week": week_label,
            "last_dates": set(),
            "positions": set(),
            "names": [],
        })
        item["last_dates"].add(record.last_date)
        item["positions"].add(record.last_position)
        item["names"].append(record.name)

    rows = []
    for _, item in sorted(buckets.items(), key=lambda entry: entry[0]):
        rows.append({
            "week": item["week"],
            "date": ", ".join(_sort_dates(item["last_dates"])),
            "position": " / ".join(sorted(item["positions"])),
            "count": len(item["names"]),
        })
    return rows


def activity_trend(records, group=GROUP_ALL):
    """날짜별 인증 인원 추이를 반환한다."""
    selected = records if group == GROUP_ALL else [r for r in records if r.group == group]
    all_dates = set()
    for record in selected:
        all_dates.update(record.activity_dates)

    rows = []
    for date_value in _sort_dates(all_dates):
        rows.append({
            "date": date_value,
            "count": sum(1 for r in selected if date_value in r.activity_dates),
        })
    return rows


def detail_rows(records):
    """상세 명단 행을 반환한다."""
    rows = []
    for record in sorted(records, key=lambda r: (r.group, r.leader, r.name)):
        rows.append([
            record.group,
            record.leader,
            record.name,
            record.emoji,
            record.progress_rate,
            record.read_count,
            record.expected_count,
            record.last_date,
            record.last_track,
            record.last_position,
            record.status,
        ])
    return rows


def _xl_text(value):
    return '"' + str(value).replace('"', '""') + '"'


def _xl_literal(value):
    if isinstance(value, (int, float)):
        return str(value)
    return _xl_text(value)


def _cell_addr(row, col):
    return f"{get_column_letter(col)}{row}"


def _range_addr(sheet_name, row, start_col, end_col):
    sheet = quote_sheetname(sheet_name)
    start = get_column_letter(start_col)
    end = get_column_letter(end_col)
    return f"{sheet}!${start}${row}:${end}${row}"


def _helper_range(helper_name, col_letter, start_row, end_row):
    sheet = quote_sheetname(helper_name)
    return f"{sheet}!${col_letter}${start_row}:${col_letter}${end_row}"


def _is_date_header(value):
    if value is None:
        return False
    month_day = _date_key(value)
    return month_day != (99, 99)


def _track_from_label(value, default_track=None):
    label = str(value or "").strip()
    if label in ("구약", "투트랙(구약)"):
        return "old"
    if label in ("신약", "투트랙(신약)"):
        return "new"
    if label == "신약일독":
        return "nt"
    if label == "성경일독":
        return "bible"
    return default_track


def _find_progress_layout(ws):
    """진도표 시트에서 헤더 행과 주요 컬럼 위치를 찾는다."""
    for row in range(1, min(ws.max_row, 12) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if "이름" not in values:
            continue
        name_col = values.index("이름") + 1
        track_col = values.index("트랙") + 1 if "트랙" in values else None
        date_cols = [
            col
            for col in range(1, ws.max_column + 1)
            if _is_date_header(ws.cell(row, col).value)
        ]
        return {
            "header_row": row,
            "name_col": name_col,
            "track_col": track_col,
            "date_start_col": min(date_cols) if date_cols else 0,
            "date_end_col": max(date_cols) if date_cols else 0,
        }
    return None


def _source_rows_for_sheet(wb, sheet_name, default_track=None):
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    layout = _find_progress_layout(ws)
    if not layout:
        return []

    rows = []
    for row in range(layout["header_row"] + 1, ws.max_row + 1):
        name = ws.cell(row, layout["name_col"]).value
        if not name:
            continue
        track = default_track
        if layout["track_col"]:
            track = _track_from_label(ws.cell(row, layout["track_col"]).value, default_track)
        rows.append((
            str(name),
            track,
            AnalysisSourceRange(
                sheet_name=sheet_name,
                header_row=layout["header_row"],
                row=row,
                date_start_col=layout["date_start_col"],
                date_end_col=layout["date_end_col"],
                track=track or "",
            ),
        ))
    return rows


def _discover_analysis_sources(wb, records):
    """분석 레코드별로 참조할 진도표 행 범위를 찾는다."""
    bible_sources = {
        name: source for name, _, source in _source_rows_for_sheet(wb, "성경일독 진도표", "bible")
    }
    nt_sources = {
        name: source for name, _, source in _source_rows_for_sheet(wb, "신약일독 진도표", "nt")
    }
    single_sources = {
        name: source for name, _, source in _source_rows_for_sheet(wb, "꿀성경 진도표")
    }

    dual_sources = {}
    for name, track, source in _source_rows_for_sheet(wb, "투트랙 진도표"):
        dual_sources.setdefault(name, {})[track] = source
    for name, _, source in _source_rows_for_sheet(wb, "구약 진도표", "old"):
        dual_sources.setdefault(name, {})["old"] = source
    for name, _, source in _source_rows_for_sheet(wb, "신약 진도표", "new"):
        dual_sources.setdefault(name, {})["new"] = source

    result = []
    for record in records:
        if record.group == GROUP_DUAL:
            by_track = dual_sources.get(record.name, {})
            sources = [by_track[track] for track in ("old", "new") if track in by_track]
        elif record.group == GROUP_NT:
            source = nt_sources.get(record.name) or single_sources.get(record.name)
            sources = [source] if source else []
            if sources and not sources[0].track:
                sources = [AnalysisSourceRange(
                    sources[0].sheet_name,
                    sources[0].header_row,
                    sources[0].row,
                    sources[0].date_start_col,
                    sources[0].date_end_col,
                    "nt",
                )]
        else:
            source = bible_sources.get(record.name) or single_sources.get(record.name)
            sources = [source] if source else []
            if sources and not sources[0].track:
                sources = [AnalysisSourceRange(
                    sources[0].sheet_name,
                    sources[0].header_row,
                    sources[0].row,
                    sources[0].date_start_col,
                    sources[0].date_end_col,
                    "bible",
                )]
        result.append(sources)
    return result


def _source_count_formula(source):
    if not source or not source.date_start_col or source.date_start_col > source.date_end_col:
        return "0"
    data_range = _range_addr(source.sheet_name, source.row, source.date_start_col, source.date_end_col)
    return f'COUNTIF({data_range},"O")'


def _source_last_formula(source):
    if not source or not source.date_start_col or source.date_start_col > source.date_end_col:
        return '""'
    data_range = _range_addr(source.sheet_name, source.row, source.date_start_col, source.date_end_col)
    header_range = _range_addr(source.sheet_name, source.header_row, source.date_start_col, source.date_end_col)
    first_data_cell = f"{quote_sheetname(source.sheet_name)}!${get_column_letter(source.date_start_col)}${source.row}"
    last_o_index = f'IFERROR(MAX(FILTER(COLUMN({data_range})-COLUMN({first_data_cell})+1,{data_range}="O")),0)'
    return f'IF({last_o_index}=0,"",INDEX({header_range},1,{last_o_index}))'


def _map_lookup_formula(date_cell, track, map_end_row, value_col, default_value):
    if not track:
        return _xl_literal(default_value)
    key_range = f"$AB$2:$AB${map_end_row}"
    value_range = f"${value_col}$2:${value_col}${map_end_row}"
    key = f"{_xl_text(track + '|')}&{date_cell}"
    default = _xl_literal(default_value)
    return f"IF({date_cell}=\"\",{default},IFERROR(INDEX({value_range},MATCH({key},{key_range},0)),{default}))"


def _record_tracks(records):
    tracks = set()
    for record in records:
        if record.group == GROUP_DUAL:
            tracks.update(("old", "new"))
        elif record.group == GROUP_NT:
            tracks.add("nt")
        else:
            tracks.add("bible")
    return tracks


def _date_order(value):
    month, day = _date_key(value)
    if month == 99:
        return 0
    return month * 100 + day


def _write_date_map(ws, records, part):
    headers = ["매핑ID", "키", "트랙", "날짜", "주", "진행 위치", "정렬값"]
    start_col = 27  # AA
    for offset, header in enumerate(headers):
        ws.cell(1, start_col + offset, header)

    row = 2
    track_order = {"bible": 0, "nt": 1, "old": 2, "new": 3}
    for track in sorted(_record_tracks(records), key=lambda value: track_order.get(value, 99)):
        for date_value in _sort_dates(expected_dates(track, part)):
            _, week_label = _week_bucket(date_value)
            ws.cell(row, start_col, row - 1)
            ws.cell(row, start_col + 1, f"{track}|{date_value}")
            ws.cell(row, start_col + 2, track)
            ws.cell(row, start_col + 3, date_value)
            ws.cell(row, start_col + 4, week_label)
            ws.cell(row, start_col + 5, _position_for_date(date_value, track, part))
            ws.cell(row, start_col + 6, _date_order(date_value))
            row += 1
    return max(2, row - 1)


def _add_formula_helper_sheet(wb, records, source_groups, part):
    if FORMULA_HELPER_SHEET in wb.sheetnames:
        del wb[FORMULA_HELPER_SHEET]
    ws = wb.create_sheet(title=FORMULA_HELPER_SHEET)
    ws.sheet_state = "hidden"

    headers = [
        "ID", "그룹", "담당", "이름", "이모티콘", "인증일수", "전체일수", "완독",
        "진행률", "마지막 인증일", "마지막 트랙", "마지막 위치", "상태",
        "소스1 마지막일", "소스1 정렬값", "소스1 위치",
        "소스2 마지막일", "소스2 정렬값", "소스2 위치", "마지막 주",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)

    map_end_row = _write_date_map(ws, records, part)

    for index, (record, sources) in enumerate(zip(records, source_groups), start=2):
        ordered_sources = sorted(sources, key=lambda source: {"old": 0, "new": 1}.get(source.track, 0))
        source1 = ordered_sources[0] if ordered_sources else None
        source2 = ordered_sources[1] if len(ordered_sources) > 1 else None

        ws.cell(index, 1, index - 1)
        ws.cell(index, 2, record.group)
        ws.cell(index, 3, record.leader)
        ws.cell(index, 4, record.name)
        ws.cell(index, 5, record.emoji)
        ws.cell(index, 6, "=" + "+".join(_source_count_formula(source) for source in ordered_sources) if ordered_sources else "=0")
        ws.cell(index, 7, record.expected_count)
        ws.cell(index, 8, f"=AND(G{index}>0,F{index}>=G{index})")
        ws.cell(index, 9, f"=IF(G{index}=0,0,MIN(1,F{index}/G{index}))")

        ws.cell(index, 14, "=" + _source_last_formula(source1))
        ws.cell(index, 15, "=" + _map_lookup_formula(f"N{index}", source1.track if source1 else "", map_end_row, "AG", 0))
        ws.cell(index, 16, "=" + _map_lookup_formula(f"N{index}", source1.track if source1 else "", map_end_row, "AF", "진도 미확인"))
        ws.cell(index, 17, "=" + _source_last_formula(source2))
        ws.cell(index, 18, "=" + _map_lookup_formula(f"Q{index}", source2.track if source2 else "", map_end_row, "AG", 0))
        ws.cell(index, 19, "=" + _map_lookup_formula(f"Q{index}", source2.track if source2 else "", map_end_row, "AF", "진도 미확인"))

        if record.group == GROUP_DUAL:
            ws.cell(index, 10, f'=IF(AND(O{index}=0,R{index}=0),"",IF(R{index}>O{index},Q{index},N{index}))')
            ws.cell(index, 11, (
                f'=IF(AND(O{index}=0,R{index}=0),"",'
                f'IF(AND(O{index}=R{index},O{index}>0),"구약/신약",'
                f'IF(R{index}>O{index},"신약","구약")))'
            ))
            ws.cell(index, 12, (
                f'=IF(J{index}="","미시작",'
                f'IF(AND(O{index}=R{index},O{index}>0),"구약: "&P{index}&" / 신약: "&S{index},'
                f'IF(R{index}>O{index},S{index},P{index})))'
            ))
        else:
            ws.cell(index, 10, f"=N{index}")
            ws.cell(index, 11, f'=IF(J{index}="","",B{index})')
            ws.cell(index, 12, f'=IF(J{index}="","미시작",P{index})')

        ws.cell(index, 13, f'=IF(H{index},"완독",IF(F{index}=0,"미시작","하차 추정"))')
        ws.cell(index, 20, (
            f'=IF(J{index}="","",IFERROR(INDEX($AE$2:$AE${map_end_row},'
            f'MATCH(IF(K{index}="신약","new",IF(K{index}="신약일독","nt",'
            f'IF(K{index}="성경일독","bible","old")))&"|"&J{index},'
            f'$AB$2:$AB${map_end_row},0)),"미확인"))'
        ))

    for col in range(1, 21):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(27, 34):
        ws.column_dimensions[get_column_letter(col)].width = 16

    return {
        "name": FORMULA_HELPER_SHEET,
        "first_row": 2,
        "last_row": max(1, len(records) + 1),
        "map_end_row": map_end_row,
    }


def _summary_formula_rows(records, helper_info, start_row, start_col):
    helper = helper_info["name"]
    first = helper_info["first_row"]
    last = helper_info["last_row"]
    if not records:
        return [[group, "=0", "=0", "=0", "=0", "=0", "=0", "=0"] for group in GROUPS]

    group_range = _helper_range(helper, "B", first, last)
    name_range = _helper_range(helper, "D", first, last)
    complete_range = _helper_range(helper, "H", first, last)
    progress_range = _helper_range(helper, "I", first, last)
    status_range = _helper_range(helper, "M", first, last)

    rows = []
    for offset, group in enumerate(GROUPS, start=1):
        excel_row = start_row + offset
        total_cell = _cell_addr(excel_row, start_col + 1)
        complete_cell = _cell_addr(excel_row, start_col + 2)
        if group == GROUP_ALL:
            total = f"=COUNTA({name_range})"
            complete = f"=COUNTIF({complete_range},TRUE)"
            avg = f"=IF({total_cell}=0,0,AVERAGE({progress_range}))"
            med = f"=IF({total_cell}=0,0,MEDIAN({progress_range}))"
            not_started = f'=COUNTIF({status_range},"미시작")'
            dropout = f'=COUNTIF({status_range},"하차 추정")'
        else:
            group_text = _xl_text(group)
            total = f"=COUNTIF({group_range},{group_text})"
            complete = f"=COUNTIFS({group_range},{group_text},{complete_range},TRUE)"
            avg = f"=IF({total_cell}=0,0,AVERAGEIF({group_range},{group_text},{progress_range}))"
            med = f"=IF({total_cell}=0,0,MEDIAN(FILTER({progress_range},{group_range}={group_text})))"
            not_started = f'=COUNTIFS({group_range},{group_text},{status_range},"미시작")'
            dropout = f'=COUNTIFS({group_range},{group_text},{status_range},"하차 추정")'
        rate = f"=IF({total_cell}=0,0,{complete_cell}/{total_cell})"
        rows.append([group, total, complete, rate, avg, med, not_started, dropout])
    return rows


def _progress_bucket_formula(helper_info, bucket, group):
    helper = helper_info["name"]
    first = helper_info["first_row"]
    last = helper_info["last_row"]
    if last < first:
        return "=0"
    group_range = _helper_range(helper, "B", first, last)
    progress_range = _helper_range(helper, "I", first, last)
    criteria = {
        "0%": [(progress_range, "<=0")],
        "1~25%": [(progress_range, ">0"), (progress_range, "<=0.25")],
        "26~50%": [(progress_range, ">0.25"), (progress_range, "<=0.5")],
        "51~75%": [(progress_range, ">0.5"), (progress_range, "<=0.75")],
        "76~99%": [(progress_range, ">0.75"), (progress_range, "<1")],
        "100%": [(progress_range, ">=1")],
    }[bucket]
    parts = []
    if group != GROUP_ALL:
        parts.extend([group_range, _xl_text(group)])
    for criteria_range, value in criteria:
        parts.extend([criteria_range, _xl_text(value)])
    return f"=COUNTIFS({','.join(parts)})"


def _formula_dropout_rows(records, helper_info, start_row, start_col):
    dropout_rows_data = dropout_week_distribution(records)
    if not dropout_rows_data:
        return [["-", "", "하차 추정 없음", ""]], [], 1

    helper = helper_info["name"]
    first = helper_info["first_row"]
    last = helper_info["last_row"]
    status_range = _helper_range(helper, "M", first, last)
    week_range = _helper_range(helper, "T", first, last)
    last_date_range = _helper_range(helper, "J", first, last)
    position_range = _helper_range(helper, "L", first, last)

    criteria_col = start_col + 11
    raw_rows = []
    for offset, item in enumerate(dropout_rows_data, start=1):
        excel_row = start_row + offset
        criteria_cell = _cell_addr(excel_row, criteria_col)
        raw_count_cell = _cell_addr(excel_row, criteria_col + 3)
        raw_count = f'COUNTIFS({status_range},"하차 추정",{week_range},{criteria_cell})'
        count_formula = f'=IF({raw_count}=0,"",{raw_count})'
        date_formula = (
            f'=IF({raw_count_cell}="","",TEXTJOIN(", ",TRUE,UNIQUE(FILTER('
            f'{last_date_range},({status_range}="하차 추정")*({week_range}={criteria_cell})))))'
        )
        position_formula = (
            f'=IF({raw_count_cell}="","",TEXTJOIN(" / ",TRUE,UNIQUE(FILTER('
            f'{position_range},({status_range}="하차 추정")*({week_range}={criteria_cell})))))'
        )
        raw_rows.append([item["week"], date_formula, position_formula, count_formula])

    raw_start = _cell_addr(start_row + 1, criteria_col)
    raw_end = _cell_addr(start_row + len(raw_rows), criteria_col + 3)
    raw_count_start = _cell_addr(start_row + 1, criteria_col + 3)
    raw_count_end = _cell_addr(start_row + len(raw_rows), criteria_col + 3)
    filtered = f"FILTER({raw_start}:{raw_end},{raw_count_start}:{raw_count_end}<>\"\")"
    formula = (
        f'=IFERROR(TAKE({filtered},{DROPOUT_VISIBLE_LIMIT}),'
        '{"-","","하차 추정 없음",""})'
    )
    visible_rows = [[formula, "", "", ""]]
    return visible_rows, raw_rows, min(DROPOUT_VISIBLE_LIMIT, max(1, len(raw_rows)))


def _source_has_date_formula(source, date_value):
    if not source or not source.date_start_col or source.date_start_col > source.date_end_col:
        return "0"
    data_range = _range_addr(source.sheet_name, source.row, source.date_start_col, source.date_end_col)
    header_range = _range_addr(source.sheet_name, source.header_row, source.date_start_col, source.date_end_col)
    return f'COUNTIFS({header_range},{_xl_text(date_value)},{data_range},"O")'


def _trend_dates_from_sources(wb, source_groups):
    dates = set()
    for sources in source_groups:
        for source in sources:
            if not source.date_start_col or source.date_start_col > source.date_end_col:
                continue
            ws = wb[source.sheet_name]
            for col in range(source.date_start_col, source.date_end_col + 1):
                value = ws.cell(source.header_row, col).value
                if _is_date_header(value):
                    dates.add(str(value))
    return _sort_dates(dates)


def _trend_count_formula(source_groups, date_value):
    terms = []
    for sources in source_groups:
        source_terms = [_source_has_date_formula(source, date_value) for source in sources]
        source_sum = "+".join(source_terms) if source_terms else "0"
        terms.append(f"--(({source_sum})>0)")
    if not terms:
        return "=0"
    return f"=SUM({','.join(terms)})"


def _formula_detail_rows(records, helper_info):
    helper = quote_sheetname(helper_info["name"])
    rows = []
    for idx, _ in enumerate(records, start=helper_info["first_row"]):
        rows.append([
            f"={helper}!B{idx}",
            f"={helper}!C{idx}",
            f"={helper}!D{idx}",
            f"={helper}!E{idx}",
            f"={helper}!I{idx}",
            f"={helper}!F{idx}",
            f"={helper}!G{idx}",
            f"={helper}!J{idx}",
            f"={helper}!K{idx}",
            f"={helper}!L{idx}",
            f"={helper}!M{idx}",
        ])
    return rows


def _visual_width(value):
    """한글이 많은 셀의 대략적인 표시 폭을 계산한다."""
    text = "" if value is None else str(value)
    return sum(2 if ord(char) > 127 else 1 for char in text)


def _estimated_line_count(value, width):
    text = "" if value is None else str(value)
    if not text:
        return 1
    if text.startswith("="):
        return 1
    capacity = max(8, int((width or 12) * 1.5))
    return sum(max(1, ceil(_visual_width(part) / capacity)) for part in text.splitlines())


def _style_table(
    ws,
    start_row,
    start_col,
    headers,
    rows,
    percent_cols=None,
    wrap_cols=None,
    left_cols=None,
):
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    percent_cols = set(percent_cols or [])
    wrap_cols = set(wrap_cols or [])
    left_cols = set(left_cols or [])

    for col_offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + col_offset, header)
        cell.fill = header_fill
        cell.font = Font(name="맑은 고딕", size=11)
        cell.alignment = center
        cell.border = border

    for row_offset, row in enumerate(rows, start=1):
        row_idx = start_row + row_offset
        max_lines = 1
        for col_offset, value in enumerate(row):
            cell = ws.cell(row_idx, start_col + col_offset, value)
            cell.font = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(
                horizontal="left" if col_offset in left_cols else "center",
                vertical="center",
                wrap_text=col_offset in wrap_cols,
            )
            cell.border = border
            if col_offset in percent_cols:
                cell.number_format = "0.0%"
            if col_offset in wrap_cols:
                col_width = ws.column_dimensions[cell.column_letter].width
                max_lines = max(max_lines, _estimated_line_count(value, col_width))
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = min(84, max(28, 18 * max_lines))


def _section_title(ws, row, col, title):
    cell = ws.cell(row, col, title)
    cell.font = Font(name="맑은 고딕", size=13, bold=True)
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(ws):
    widths = {
        "A": 2,
        "B": 16,
        "C": 16,
        "D": 24,
        "E": 13,
        "F": 42,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 18,
        "K": 32,
        "L": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 6


def add_analysis_sheet(wb, records):
    """Workbook에 분석결과 시트를 추가한다."""
    ws = wb.create_sheet(title="분석결과")
    _set_widths(ws)

    R = ROW_PAD
    C = COL_PAD
    row = 1 + R
    col = 1 + C

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 10)
    title_cell = ws.cell(row, col, "분석결과")
    title_cell.font = Font(name="맑은 고딕", size=20)
    title_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 42

    row += 2
    _section_title(ws, row, col, "요약")
    summary_headers = ["그룹", "전체 인원", "완독자", "완독률", "평균 진행률", "중앙 진행률", "미시작", "하차 추정"]
    summary_rows = [
        [
            item["group"],
            item["total"],
            item["complete"],
            item["complete_rate"],
            item["avg_progress"],
            item["median_progress"],
            item["not_started"],
            item["dropout"],
        ]
        for item in summarize_records(records)
    ]
    _style_table(ws, row + 1, col, summary_headers, summary_rows, percent_cols={3, 4, 5})

    row += len(summary_rows) + 4
    _section_title(ws, row, col, "진행률 구간 분포")
    distribution = progress_distribution(records)
    dist_headers = ["구간"] + GROUPS
    dist_rows = [[item["bucket"]] + [item[group] for group in GROUPS] for item in distribution]
    dist_table_row = row + 1
    _style_table(ws, dist_table_row, col, dist_headers, dist_rows)

    chart = BarChart()
    chart.title = "진행률 구간 분포"
    chart.y_axis.title = "인원"
    chart.x_axis.title = "진행률"
    data = Reference(ws, min_col=col + 1, max_col=col + len(GROUPS),
                     min_row=dist_table_row, max_row=dist_table_row + len(dist_rows))
    categories = Reference(ws, min_col=col, min_row=dist_table_row + 1,
                           max_row=dist_table_row + len(dist_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 16
    ws.add_chart(chart, "H11")

    row += len(dist_rows) + 4
    _section_title(ws, row, col, "하차 추정 분포")
    dropout_rows_data = dropout_week_distribution(records)
    dropout_headers = ["하차 주", "마지막 인증일", "진행 위치", "하차 추정 인원"]
    dropout_rows = [
        [item["week"], item["date"], item["position"], item["count"]]
        for item in dropout_rows_data
    ]
    if not dropout_rows:
        dropout_rows = [["-", "-", "하차 추정 없음", 0]]
    dropout_table_row = row + 1
    _style_table(
        ws,
        dropout_table_row,
        col,
        dropout_headers,
        dropout_rows,
        wrap_cols={1, 2},
        left_cols={2},
    )

    if any(row_data[3] for row_data in dropout_rows):
        chart = BarChart()
        chart.title = "하차 추정 분포"
        chart.y_axis.title = "인원"
        chart.x_axis.title = "하차 주"
        data = Reference(ws, min_col=col + 3, min_row=dropout_table_row,
                         max_row=dropout_table_row + len(dropout_rows))
        categories = Reference(ws, min_col=col, min_row=dropout_table_row + 1,
                               max_row=dropout_table_row + len(dropout_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "H25")

    row += len(dropout_rows) + 4
    _section_title(ws, row, col, "날짜별 인증 인원 추이")
    trend = activity_trend(records)
    trend_headers = ["날짜", "인증 인원"]
    trend_rows = [[item["date"], item["count"]] for item in trend] or [["-", 0]]
    trend_table_row = row + 1
    _style_table(ws, trend_table_row, col, trend_headers, trend_rows)

    if any(row_data[1] for row_data in trend_rows):
        chart = LineChart()
        chart.title = "날짜별 인증 인원 추이"
        chart.y_axis.title = "인원"
        chart.x_axis.title = "날짜"
        data = Reference(ws, min_col=col + 1, min_row=trend_table_row,
                         max_row=trend_table_row + len(trend_rows))
        categories = Reference(ws, min_col=col, min_row=trend_table_row + 1,
                               max_row=trend_table_row + len(trend_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "H39")

    row += len(trend_rows) + 4
    _section_title(ws, row, col, "상세 명단")
    detail_headers = [
        "그룹", "담당", "이름", "이모티콘", "진행률", "인증일수", "전체일수",
        "마지막 인증일", "마지막 트랙", "마지막 위치", "상태",
    ]
    _style_table(
        ws,
        row + 1,
        col,
        detail_headers,
        detail_rows(records),
        percent_cols={4},
        wrap_cols={1, 2, 9},
        left_cols={1, 2, 9},
    )

    ws.freeze_panes = "B3"
    return ws


def add_formula_analysis_sheet(wb, records, part=1):
    """Workbook에 진도표를 참조하는 수식 기반 분석결과 시트를 추가한다."""
    source_groups = _discover_analysis_sources(wb, records)
    helper_info = _add_formula_helper_sheet(wb, records, source_groups, normalize_part(part))

    ws = wb.create_sheet(title="분석결과")
    _set_widths(ws)

    # 파일을 열 때 Excel/Sheets가 helper 수식을 다시 계산하도록 힌트를 남긴다.
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    R = ROW_PAD
    C = COL_PAD
    row = 1 + R
    col = 1 + C

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 10)
    title_cell = ws.cell(row, col, "분석결과")
    title_cell.font = Font(name="맑은 고딕", size=20)
    title_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 42

    row += 2
    _section_title(ws, row, col, "요약")
    summary_headers = ["그룹", "전체 인원", "완독자", "완독률", "평균 진행률", "중앙 진행률", "미시작", "하차 추정"]
    summary_table_row = row + 1
    summary_rows = _summary_formula_rows(records, helper_info, summary_table_row, col)
    _style_table(ws, summary_table_row, col, summary_headers, summary_rows, percent_cols={3, 4, 5})

    row += len(summary_rows) + 4
    _section_title(ws, row, col, "진행률 구간 분포")
    dist_headers = ["구간"] + GROUPS
    dist_rows = [
        [bucket] + [_progress_bucket_formula(helper_info, bucket, group) for group in GROUPS]
        for bucket in PROGRESS_BUCKETS
    ]
    dist_table_row = row + 1
    _style_table(ws, dist_table_row, col, dist_headers, dist_rows)

    chart = BarChart()
    chart.title = "진행률 구간 분포"
    chart.y_axis.title = "인원"
    chart.x_axis.title = "진행률"
    data = Reference(ws, min_col=col + 1, max_col=col + len(GROUPS),
                     min_row=dist_table_row, max_row=dist_table_row + len(dist_rows))
    categories = Reference(ws, min_col=col, min_row=dist_table_row + 1,
                           max_row=dist_table_row + len(dist_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 16
    ws.add_chart(chart, "H11")

    row += len(dist_rows) + 4
    _section_title(ws, row, col, "하차 추정 분포")
    dropout_headers = ["하차 주", "마지막 인증일", "진행 위치", "하차 추정 인원"]
    dropout_table_row = row + 1
    dropout_rows, dropout_raw_rows, dropout_visible_slots = _formula_dropout_rows(
        records,
        helper_info,
        dropout_table_row,
        col,
    )
    _style_table(
        ws,
        dropout_table_row,
        col,
        dropout_headers,
        dropout_rows,
        wrap_cols={1, 2},
        left_cols={2},
    )
    criteria_col = col + 11
    for row_offset, raw_row in enumerate(dropout_raw_rows, start=1):
        for col_offset, value in enumerate(raw_row):
            ws.cell(dropout_table_row + row_offset, criteria_col + col_offset, value)
    for hidden_col in range(criteria_col, criteria_col + 4):
        ws.column_dimensions[get_column_letter(hidden_col)].hidden = True

    chart = BarChart()
    chart.title = "하차 추정 분포"
    chart.y_axis.title = "인원"
    chart.x_axis.title = "하차 주"
    data = Reference(ws, min_col=col + 3, min_row=dropout_table_row,
                     max_row=dropout_table_row + dropout_visible_slots)
    categories = Reference(ws, min_col=col, min_row=dropout_table_row + 1,
                           max_row=dropout_table_row + dropout_visible_slots)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 16
    ws.add_chart(chart, "H25")

    row += dropout_visible_slots + 4
    _section_title(ws, row, col, "날짜별 인증 인원 추이")
    trend_headers = ["날짜", "인증 인원"]
    trend_dates = _trend_dates_from_sources(wb, source_groups)
    trend_rows = [[date_value, _trend_count_formula(source_groups, date_value)] for date_value in trend_dates]
    if not trend_rows:
        trend_rows = [["-", "=0"]]
    trend_table_row = row + 1
    _style_table(ws, trend_table_row, col, trend_headers, trend_rows)

    chart = LineChart()
    chart.title = "날짜별 인증 인원 추이"
    chart.y_axis.title = "인원"
    chart.x_axis.title = "날짜"
    data = Reference(ws, min_col=col + 1, min_row=trend_table_row,
                     max_row=trend_table_row + len(trend_rows))
    categories = Reference(ws, min_col=col, min_row=trend_table_row + 1,
                           max_row=trend_table_row + len(trend_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 16
    ws.add_chart(chart, "H39")

    row += len(trend_rows) + 4
    _section_title(ws, row, col, "상세 명단")
    detail_headers = [
        "그룹", "담당", "이름", "이모티콘", "진행률", "인증일수", "전체일수",
        "마지막 인증일", "마지막 트랙", "마지막 위치", "상태",
    ]
    _style_table(
        ws,
        row + 1,
        col,
        detail_headers,
        _formula_detail_rows(records, helper_info),
        percent_cols={4},
        wrap_cols={1, 2, 9},
        left_cols={1, 2, 9},
    )

    ws.freeze_panes = "B3"
    return ws
