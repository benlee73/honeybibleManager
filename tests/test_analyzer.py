import csv
import io

import pytest

from app.analyzer import (
    analyze_chat,
    build_dual_preview_data,
    build_output_csv,
    build_output_xlsx,
    build_preview_data,
    choose_assigned_emoji,
    decode_csv_payload,
    decode_payload,
    extract_tracks,
    iter_data_rows,
    message_contains_emoji,
    normalize_user_name,
    parse_csv_rows,
    sort_dates,
)


class TestDecodeCsvPayload:
    def test_decode_csv_payload__utf8_bom__decodes_correctly(self):
        text = "한글 텍스트"
        payload = b"\xef\xbb\xbf" + text.encode("utf-8")
        assert decode_csv_payload(payload) == text

    def test_decode_csv_payload__utf8__decodes_correctly(self):
        text = "hello world"
        payload = text.encode("utf-8")
        assert decode_csv_payload(payload) == text

    def test_decode_csv_payload__cp949__decodes_correctly(self):
        text = "한글 텍스트"
        payload = text.encode("cp949")
        assert decode_csv_payload(payload) == text

    def test_decode_csv_payload__invalid_bytes__replaces_errors(self):
        payload = b"\xff\xfe\xfd\x80\x81"
        result = decode_csv_payload(payload)
        assert isinstance(result, str)
        assert "\ufffd" in result


class TestIterDataRows:
    def test_iter_data_rows__with_header__skips_header(self):
        csv_text = "날짜,이름,메시지\n2024-01-01,user1,hello"
        reader = csv.reader(io.StringIO(csv_text, newline=""))
        rows = list(iter_data_rows(reader))
        assert len(rows) == 1
        assert rows[0][1] == "user1"

    def test_iter_data_rows__data_starts_with_date__yields_first_row(self):
        csv_text = "2024-01-01,user1,hello\n2024-01-02,user2,world"
        reader = csv.reader(io.StringIO(csv_text, newline=""))
        rows = list(iter_data_rows(reader))
        assert len(rows) == 2

    def test_iter_data_rows__empty_reader__yields_nothing(self):
        reader = csv.reader(io.StringIO("", newline=""))
        rows = list(iter_data_rows(reader))
        assert rows == []


class TestChooseAssignedEmoji:
    def test_choose_assigned_emoji__highest_count__returns_it(self):
        counts = {"😀": 3, "🔥": 1}
        order = ["😀", "🔥"]
        assert choose_assigned_emoji(counts, order) == "😀"

    def test_choose_assigned_emoji__tie__returns_first_in_order(self):
        counts = {"😀": 2, "🔥": 2}
        order = ["🔥", "😀"]
        assert choose_assigned_emoji(counts, order) == "🔥"

    def test_choose_assigned_emoji__empty_counts__returns_empty_string(self):
        assert choose_assigned_emoji({}, []) == ""

    def test_choose_assigned_emoji__max_not_in_order__returns_first_key(self):
        counts = {"😀": 3, "🔥": 1}
        order = ["🎉"]
        result = choose_assigned_emoji(counts, order)
        assert result == "😀"


class TestMessageContainsEmoji:
    def test_message_contains_emoji__raw_emoji_present__returns_true(self):
        assert message_contains_emoji("hello😀world", "😀", "😀") is True

    def test_message_contains_emoji__normalized_match__returns_true(self):
        assert message_contains_emoji("hello☀\uFE0Fworld", "☀", None) is True

    def test_message_contains_emoji__not_present__returns_false(self):
        assert message_contains_emoji("hello world", "😀", "😀") is False

    def test_message_contains_emoji__empty_raw__falls_through_to_normalized(self):
        assert message_contains_emoji("hello😀world", "😀", "") is True


class TestAnalyzeChat:
    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_analyze_chat__normal_csv__assigns_emoji_and_collects_dates(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/15😀"],
            ["2024-01-02", "user1", "3/16😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "😀" == result["user1"]["emoji"]
        assert "3/15" in result["user1"]["dates"]
        assert "3/16" in result["user1"]["dates"]

    def test_analyze_chat__user_without_emoji__excluded(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/15"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" not in result

    def test_analyze_chat__empty_csv__returns_empty(self):
        result = analyze_chat("")
        assert result == {}

    def test_analyze_chat__message_without_dates__not_counted(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "안녕하세요😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" not in result

    def test_analyze_chat__날짜_수_상한_초과_메시지__스킵(self):
        # 1/1~2/28 범위는 59개 날짜로 확장되어 상한(30) 초과 → 스킵
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/15😀"],
            ["2024-01-02", "user1", "1/1~2/28😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        # 상한 초과 메시지의 날짜는 포함되지 않아야 함
        assert result["user1"]["dates"] == {"3/15"}

    def test_analyze_chat__날짜_수_상한_이하_메시지__정상_처리(self):
        # 2/2~2/20 범위는 19개 날짜로 상한(30) 이하 → 정상 처리
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2~2/20😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "2/2" in result["user1"]["dates"]
        assert "2/10" in result["user1"]["dates"]

    def test_analyze_chat__같은_날짜_여러_메시지__1회로_카운팅(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/15😀"],
            ["2024-01-02", "user1", "3/15😀"],
            ["2024-01-03", "user1", "3/15😀"],
            ["2024-01-04", "user1", "3/16😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"3/15", "3/16"}
        assert len(result["user1"]["dates"]) == 2


class TestSortDates:
    def test_sort_dates__unordered__sorts_by_month_then_day(self):
        dates = ["3/15", "1/5", "3/1", "2/28"]
        assert sort_dates(dates) == ["1/5", "2/28", "3/1", "3/15"]

    def test_sort_dates__empty__returns_empty(self):
        assert sort_dates([]) == []

    def test_sort_dates__single_element__returns_same(self):
        assert sort_dates(["5/10"]) == ["5/10"]


class TestBuildOutputCsv:
    def test_build_output_csv__normal_users__produces_valid_csv(self):
        users = {
            "user1": {"dates": {"3/15", "3/16"}, "emoji": "😀"},
            "user2": {"dates": {"1/5"}, "emoji": "🔥"},
        }
        output = build_output_csv(users)
        assert output[:3] == b"\xef\xbb\xbf"
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        header = rows[0]
        assert header[0] == "이름"
        assert header[1] == "이모티콘"
        assert header[2:] == ["1/5", "3/15", "3/16"]
        assert len(rows) == 3

    def test_build_output_csv__date_columns_marked_with_o(self):
        users = {
            "user1": {"dates": {"2/2", "2/3", "2/5"}, "emoji": "😀"},
            "user2": {"dates": {"2/2", "2/4"}, "emoji": "🔥"},
        }
        output = build_output_csv(users)
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        header = rows[0]
        assert header[2:] == ["2/2", "2/3", "2/4", "2/5"]
        # user1: 2/2 O, 2/3 O, 2/4 빈칸, 2/5 O
        user1_row = rows[1]
        assert user1_row[0] == "user1"
        assert user1_row[2:] == ["O", "O", "", "O"]
        # user2: 2/2 O, 2/3 빈칸, 2/4 O, 2/5 빈칸
        user2_row = rows[2]
        assert user2_row[0] == "user2"
        assert user2_row[2:] == ["O", "", "O", ""]

    def test_build_output_csv__user_with_no_dates__빈_행으로_포함(self):
        users = {
            "user1": {"dates": set(), "emoji": "😀"},
            "user2": {"dates": {"3/15"}, "emoji": "🔥"},
        }
        output = build_output_csv(users)
        text = output.decode("utf-8-sig")
        assert "user1" in text
        assert "user2" in text

    def test_build_output_csv__bom_encoding__present(self):
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        output = build_output_csv(users)
        assert output.startswith(b"\xef\xbb\xbf")

    def test_build_output_csv__empty_users__header_only(self):
        output = build_output_csv({})
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0][0] == "이름"


class TestExtractTracks:
    def test_extract_tracks__구약만__old_반환(self):
        assert extract_tracks("2/2 구약 🐷") == {"old"}

    def test_extract_tracks__신약만__new_반환(self):
        assert extract_tracks("2/2 신약 🐷") == {"new"}

    def test_extract_tracks__둘_다__old_new_반환(self):
        assert extract_tracks("2/2 구약 신약 🐷") == {"old", "new"}

    def test_extract_tracks__키워드_없음__양쪽_트랙_반환(self):
        assert extract_tracks("2/2 🐷") == {"old", "new"}


class TestAnalyzeChatDual:
    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_analyze_chat_dual__구약_신약_날짜_분리(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-02", "user1", "2/3 신약 😀"],
            ["2024-01-03", "user1", "2/4 구약 신약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2", "2/4"}
        assert result["user1"]["dates_new"] == {"2/3", "2/4"}

    def test_analyze_chat_dual__키워드_없는_메시지_양쪽_트랙_체크(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-02", "user1", "2/3 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2", "2/3"}
        assert result["user1"]["dates_new"] == {"2/3"}

    def test_analyze_chat_dual__키워드_없는_범위_메시지_양쪽_트랙_체크(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-02", "user1", "~2/4 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        # ~2/4는 구약 last_old=2/2 기준으로 2/3,2/4 확장, 신약은 last_new=None이므로 2/4만
        assert "2/3" in result["user1"]["dates_old"]
        assert "2/4" in result["user1"]["dates_old"]
        assert "2/4" in result["user1"]["dates_new"]

    def test_analyze_chat_dual__키워드_있는_메시지와_없는_메시지_혼합(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-01", "user1", "2/2 신약 😀"],
            ["2024-01-02", "user1", "2/3 😀"],
            ["2024-01-03", "user1", "2/4 구약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        # 2/3은 키워드 없으므로 양쪽에 체크
        assert result["user1"]["dates_old"] == {"2/2", "2/3", "2/4"}
        assert result["user1"]["dates_new"] == {"2/2", "2/3"}

    def test_analyze_chat_dual__여러_사용자(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-01", "user2", "2/2 신약 🔥"],
            ["2024-01-02", "user2", "2/3 구약 신약 🔥"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert result["user1"]["dates_old"] == {"2/2"}
        assert result["user1"]["dates_new"] == set()
        assert result["user2"]["dates_old"] == {"2/3"}
        assert result["user2"]["dates_new"] == {"2/2", "2/3"}

    def test_analyze_chat_dual__날짜_수_상한_초과_메시지__스킵(self):
        # 1/1~2/28 범위는 59개 날짜로 확장되어 상한(30) 초과 → 스킵
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/14 구약 😀"],
            ["2024-01-02", "user1", "1/1~2/28 구약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        # 상한 초과 메시지의 날짜는 포함되지 않아야 함
        assert result["user1"]["dates_old"] == {"3/14"}

    def test_analyze_chat_dual__같은_날짜_여러_메시지__1회로_카운팅(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-02", "user1", "2/2 구약 😀"],
            ["2024-01-03", "user1", "2/3 신약 😀"],
            ["2024-01-04", "user1", "2/3 신약 😀"],
            ["2024-01-05", "user1", "2/4 구약 신약 😀"],
            ["2024-01-06", "user1", "2/4 구약 신약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2", "2/4"}
        assert result["user1"]["dates_new"] == {"2/3", "2/4"}
        assert len(result["user1"]["dates_old"]) == 2
        assert len(result["user1"]["dates_new"]) == 2

    def test_analyze_chat_dual__텍스트_이모티콘_사용자__정상_분석(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "광천 강창우", "2/2 구약 신약 (무표정)"],
            ["2024-01-02", "광천 강창우", "2/4 구약 신약 (무표정)"],
            ["2024-01-03", "광천 김형은", "2/2~3 구약 신약(연필)"],
            ["2024-01-04", "광천 김형은", "2/3~6 구약 신약(연필)"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        # "광천 강창우" → 정규화 → "강창우"
        assert "강창우" in result
        assert result["강창우"]["emoji"] == "(무표정)"
        assert result["강창우"]["dates_old"] == {"2/2", "2/4"}
        assert result["강창우"]["dates_new"] == {"2/2", "2/4"}

        # "광천 김형은" → 정규화 → "김형은"
        assert "김형은" in result
        assert result["김형은"]["emoji"] == "(연필)"
        assert result["김형은"]["dates_old"] == {"2/2", "2/3", "2/4", "2/5", "2/6"}
        assert result["김형은"]["dates_new"] == {"2/2", "2/3", "2/4", "2/5", "2/6"}

    def test_analyze_chat_single__기존_동작_유지(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2😀"],
        ])
        result = analyze_chat(csv_text, track_mode="single")
        assert "dates" in result["user1"]
        assert result["user1"]["dates"] == {"2/2"}


class TestBuildOutputCsvDual:
    def test_build_output_csv_dual__트랙_컬럼_포함(self):
        users = {
            "user1": {"dates_old": {"2/2", "2/4"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        output = build_output_csv(users, track_mode="dual")
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        header = rows[0]
        assert header[0] == "이름"
        assert header[1] == "이모티콘"
        assert header[2] == "트랙"

    def test_build_output_csv_dual__사용자별_구약_신약_묶음(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
            "user2": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "🔥"},
        }
        output = build_output_csv(users, track_mode="dual")
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        # 헤더 + user1구약 + user1신약 + user2구약 + user2신약 = 5행
        assert len(rows) == 5
        assert rows[1][0] == "user1"
        assert rows[1][2] == "구약"
        assert rows[2][0] == "user1"
        assert rows[2][2] == "신약"
        assert rows[3][0] == "user2"
        assert rows[3][2] == "구약"
        assert rows[4][0] == "user2"
        assert rows[4][2] == "신약"

    def test_build_output_csv_dual__빈_트랙_사용자_생략(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": set(), "emoji": "😀"},
            "user2": {"dates_old": set(), "dates_new": {"2/3"}, "emoji": "🔥"},
        }
        output = build_output_csv(users, track_mode="dual")
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        # 헤더 + user1구약 + user2신약 = 3행
        assert len(rows) == 3
        assert rows[1][0] == "user1"
        assert rows[1][2] == "구약"
        assert rows[2][0] == "user2"
        assert rows[2][2] == "신약"

    def test_build_output_csv_dual__날짜_마킹_정확(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        output = build_output_csv(users, track_mode="dual")
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        header = rows[0]
        date_cols = header[3:]
        assert date_cols == ["2/2", "2/3"]
        # 구약 행: 2/2=O, 2/3=빈칸
        old_row = rows[1]
        assert old_row[3:] == ["O", ""]
        # 신약 행: 2/2=빈칸, 2/3=O
        new_row = rows[2]
        assert new_row[3:] == ["", "O"]

    def test_build_output_csv_dual__빈_사용자__헤더만(self):
        output = build_output_csv({}, track_mode="dual")
        text = output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text, newline=""))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0][:3] == ["이름", "이모티콘", "트랙"]


class TestBuildPreviewData:
    def test_build_preview_data__single__헤더와_행_구조(self):
        users = {
            "user1": {"dates": {"3/15", "3/16"}, "emoji": "😀"},
            "user2": {"dates": {"1/5"}, "emoji": "🔥"},
        }
        headers, rows = build_preview_data(users)
        assert headers[0] == "이름"
        assert headers[1] == "이모티콘"
        assert headers[2:] == ["1/5", "3/15", "3/16"]
        assert len(rows) == 2

    def test_build_preview_data__dual__트랙_컬럼_포함(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        headers, rows = build_preview_data(users, track_mode="dual")
        assert headers[:3] == ["이름", "이모티콘", "트랙"]
        assert len(rows) == 2
        assert rows[0][2] == "구약"
        assert rows[1][2] == "신약"

    def test_build_preview_data__빈_사용자__행_없음(self):
        headers, rows = build_preview_data({})
        assert headers == ["이름", "이모티콘"]
        assert rows == []

    def test_build_preview_data__빈_날짜_사용자_포함(self):
        users = {
            "user1": {"dates": set(), "emoji": "😀"},
            "user2": {"dates": {"3/15"}, "emoji": "🔥"},
        }
        headers, rows = build_preview_data(users)
        assert len(rows) == 2
        user_names = [r[0] for r in rows]
        assert "user1" in user_names
        assert "user2" in user_names

    def test_build_preview_data__O_마크_정확(self):
        users = {
            "user1": {"dates": {"2/2", "2/4"}, "emoji": "😀"},
        }
        headers, rows = build_preview_data(users)
        assert headers[2:] == ["2/2", "2/4"]
        assert rows[0][2:] == ["O", "O"]

    def test_build_preview_data__csv와_동일_데이터_single(self):
        users = {
            "user1": {"dates": {"3/15", "3/16"}, "emoji": "😀"},
            "user2": {"dates": {"1/5"}, "emoji": "🔥"},
        }
        headers, rows = build_preview_data(users)
        csv_output = build_output_csv(users)
        csv_text = csv_output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(csv_text, newline=""))
        csv_rows = list(reader)
        assert headers == csv_rows[0]
        assert rows == csv_rows[1:]

    def test_build_preview_data__csv와_동일_데이터_dual(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        headers, rows = build_preview_data(users, track_mode="dual")
        csv_output = build_output_csv(users, track_mode="dual")
        csv_text = csv_output.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(csv_text, newline=""))
        csv_rows = list(reader)
        assert headers == csv_rows[0]
        assert rows == csv_rows[1:]


class TestBuildOutputXlsx:
    def test_build_output_xlsx__반환값이_bytes(self):
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_build_output_xlsx__유효한_xlsx_파일(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.title == "꿀성경 진도표"

    def test_build_output_xlsx__헤더_행_내용(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(2, 2).value == "이름"
        assert ws.cell(2, 3).value == "이모티콘"
        assert ws.cell(2, 4).value == "2/2"
        assert ws.cell(2, 5).value == "2/3"

    def test_build_output_xlsx__데이터_행_O_마크(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(3, 2).value == "user1"
        assert ws.cell(3, 3).value == "😀"
        assert ws.cell(3, 4).value == "O"

    def test_build_output_xlsx__헤더_스타일(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        cell = ws.cell(2, 2)
        assert cell.fill.start_color.rgb == "00D6E4F0"

    def test_build_output_xlsx__O_마크_폰트_스타일(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        mark_cell = ws.cell(3, 4)
        assert mark_cell.value == "O"

    def test_build_output_xlsx__고정_틀(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.freeze_panes == "B3"

    def test_build_output_xlsx__dual_모드__시트_2개_생성(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2

    def test_build_output_xlsx__dual_모드__시트_이름(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        assert wb.sheetnames == ["구약 진도표", "신약 진도표"]

    def test_build_output_xlsx__dual_모드__트랙_컬럼_없음(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_old = wb["구약 진도표"]
        ws_new = wb["신약 진도표"]
        # 헤더에 "트랙" 컬럼이 없어야 함
        old_headers = [ws_old.cell(2, c).value for c in range(2, ws_old.max_column + 1)]
        new_headers = [ws_new.cell(2, c).value for c in range(2, ws_new.max_column + 1)]
        assert "트랙" not in old_headers
        assert "트랙" not in new_headers

    def test_build_output_xlsx__dual_모드__구약_시트_데이터(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_old = wb["구약 진도표"]
        assert ws_old.cell(2, 2).value == "이름"
        assert ws_old.cell(2, 3).value == "이모티콘"
        assert ws_old.cell(2, 4).value == "2/2"
        assert ws_old.cell(3, 2).value == "user1"
        assert ws_old.cell(3, 4).value == "O"

    def test_build_output_xlsx__dual_모드__신약_시트_데이터(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_new = wb["신약 진도표"]
        assert ws_new.cell(2, 2).value == "이름"
        assert ws_new.cell(2, 3).value == "이모티콘"
        assert ws_new.cell(2, 4).value == "2/3"
        assert ws_new.cell(3, 2).value == "user1"
        assert ws_new.cell(3, 4).value == "O"

    def test_build_output_xlsx__dual_모드__한쪽_트랙만_있는_사용자(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": set(), "emoji": "😀"},
            "user2": {"dates_old": set(), "dates_new": {"2/3"}, "emoji": "🔥"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_old = wb["구약 진도표"]
        ws_new = wb["신약 진도표"]
        # 구약 시트에는 user1만
        assert ws_old.cell(3, 2).value == "user1"
        assert ws_old.cell(4, 2).value is None
        # 신약 시트에는 user2만
        assert ws_new.cell(3, 2).value == "user2"
        assert ws_new.cell(4, 2).value is None

    def test_build_output_xlsx__dual_모드__각_시트_스타일_적용(self):
        from openpyxl import load_workbook
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        for sheet_name in ["구약 진도표", "신약 진도표"]:
            ws = wb[sheet_name]
            # 헤더 스타일
            assert ws.cell(2, 2).fill.start_color.rgb == "00D6E4F0"
            # O 마크 스타일
            assert ws.cell(3, 4).value == "O"
            # 고정 틀
            assert ws.freeze_panes == "B3"

    def test_build_output_xlsx__dual_모드__각_시트_날짜_독립(self):
        from openpyxl import load_workbook
        users = {
            "user1": {
                "dates_old": {"2/2", "2/4"},
                "dates_new": {"2/3", "2/5"},
                "emoji": "😀",
            },
        }
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_old = wb["구약 진도표"]
        ws_new = wb["신약 진도표"]
        # 구약 시트 날짜 컬럼: 2/2, 2/4만
        old_dates = [ws_old.cell(2, c).value for c in range(4, ws_old.max_column + 1)]
        assert old_dates == ["2/2", "2/4"]
        # 신약 시트 날짜 컬럼: 2/3, 2/5만
        new_dates = [ws_new.cell(2, c).value for c in range(4, ws_new.max_column + 1)]
        assert new_dates == ["2/3", "2/5"]

    def test_build_output_xlsx__빈_사용자__헤더만(self):
        from openpyxl import load_workbook
        result = build_output_xlsx({})
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(2, 2).value == "이름"
        assert ws.cell(3, 2).value is None


class TestMetaSheet:
    def test_meta_시트_생성_및_숨김(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        meta = {"room_name": "테스트방", "track_mode": "single", "schedule_type": "bible", "leader": "방장"}
        result = build_output_xlsx(users, meta=meta)
        wb = load_workbook(io.BytesIO(result))
        assert "_메타" in wb.sheetnames
        ws = wb["_메타"]
        assert ws.sheet_state == "hidden"

    def test_meta_시트_내용(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        meta = {"room_name": "꿀성경 교육국", "track_mode": "single", "schedule_type": "education", "leader": "길동"}
        result = build_output_xlsx(users, meta=meta)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["_메타"]
        meta_dict = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] is not None:
                meta_dict[row[0]] = row[1]
        assert meta_dict["room_name"] == "꿀성경 교육국"
        assert meta_dict["track_mode"] == "single"
        assert meta_dict["schedule_type"] == "education"
        assert meta_dict["leader"] == "길동"

    def test_meta_None__메타_시트_없음(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users)
        wb = load_workbook(io.BytesIO(result))
        assert "_메타" not in wb.sheetnames

    def test_meta_빈_dict__메타_시트_없음(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates": {"3/15"}, "emoji": "😀"}}
        result = build_output_xlsx(users, meta={})
        wb = load_workbook(io.BytesIO(result))
        assert "_메타" not in wb.sheetnames

    def test_meta_dual_모드__메타_시트_포함(self):
        from openpyxl import load_workbook
        users = {"user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"}}
        meta = {"room_name": "투트랙방", "track_mode": "dual", "schedule_type": "dual", "leader": "방장"}
        result = build_output_xlsx(users, track_mode="dual", meta=meta)
        wb = load_workbook(io.BytesIO(result))
        assert "_메타" in wb.sheetnames
        assert "구약 진도표" in wb.sheetnames
        assert "신약 진도표" in wb.sheetnames


class TestAnalyzeChatLeadingTildeCatchup:
    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_single_모드_캐치업__이전_인증_후_선행_틸드(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/4😀"],
            ["2024-01-02", "user1", "~2/7😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"2/4", "2/5", "2/6", "2/7"}

    def test_첫_메시지_선행_틸드__last_date_없음__단일_날짜(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "~2/7😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"2/7"}

    def test_30일_상한_초과__스킵(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "1/1😀"],
            ["2024-01-02", "user1", "~2/28😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        # ~2/28은 1/2~2/28=58일 → 상한(30) 초과 → 스킵
        assert result["user1"]["dates"] == {"1/1"}

    def test_dual_모드__트랙별_last_date_독립_추적(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/2 구약 😀"],
            ["2024-01-02", "user1", "2/4 신약 😀"],
            ["2024-01-03", "user1", "~2/5 구약 😀"],
            ["2024-01-04", "user1", "~2/6 신약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        # 구약 last_date=(2,2) → ~2/5 → 2/3,2/4,2/5
        assert result["user1"]["dates_old"] == {"2/2", "2/3", "2/4", "2/5"}
        # 신약 last_date=(2,4) → ~2/6 → 2/5,2/6
        assert result["user1"]["dates_new"] == {"2/4", "2/5", "2/6"}

    def test_연속_캐치업(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/1😀"],
            ["2024-01-02", "user1", "~2/3😀"],
            ["2024-01-03", "user1", "~2/5😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"2/1", "2/2", "2/3", "2/4", "2/5"}


class TestAnalyzeChatScheduleFilter:
    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_single_성경일독_키워드__일요일_날짜_제외(self):
        # 2/8 = 일요일 → 진도표에 없으므로 제외
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "창세기 1장 2/7😀"],
            ["2024-01-02", "user1", "출애굽기 2장 2/8😀"],
            ["2024-01-03", "user1", "2/9😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "2/7" in result["user1"]["dates"]
        assert "2/8" not in result["user1"]["dates"]
        assert "2/9" in result["user1"]["dates"]

    def test_single_키워드_없음__필터링_없이_전체_통과(self):
        # 키워드 없으면 schedule=None → 필터 미적용
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "2/8😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "2/8" in result["user1"]["dates"]

    def test_dual_모드__트랙별_진도표_적용(self):
        # dual 모드: 구약 → BIBLE_DATES, 신약 → NT_DATES
        # 5/30: 성경일독O, 신약일독X
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "5/30 구약 😀"],
            ["2024-01-02", "user1", "5/30 신약 😀"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        assert "5/30" in result["user1"]["dates_old"]
        assert "5/30" not in result["user1"]["dates_new"]

    def test_선행_틸드_확장_후_필터_적용(self):
        # 2/7 토→2/9 월, 2/8=일요일 제외
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "창세기 2/7😀"],
            ["2024-01-02", "user1", "출애굽기 ~2/9😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "2/7" in result["user1"]["dates"]
        assert "2/8" not in result["user1"]["dates"]
        assert "2/9" in result["user1"]["dates"]

    def test_진도표_외_날짜만_포함된_메시지__해당_메시지_날짜_없음(self):
        # 일요일 날짜만 있는 메시지 → 필터 후 빈 리스트 → skip
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "창세기 2/2😀"],
            ["2024-01-02", "user1", "출애굽기 2/3😀"],
            ["2024-01-03", "user1", "2/8😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "2/2" in result["user1"]["dates"]
        assert "2/3" in result["user1"]["dates"]
        assert "2/8" not in result["user1"]["dates"]


class TestParseCsvRows:
    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_parse_csv_rows__헤더_스킵_후_데이터_반환(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "hello"],
        ])
        rows = parse_csv_rows(csv_text)
        assert len(rows) == 1
        assert rows[0] == ("user1", "hello")

    def test_parse_csv_rows__빈_입력__빈_리스트(self):
        assert parse_csv_rows("") == []

    def test_parse_csv_rows__짧은_행_스킵(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1"],
        ])
        rows = parse_csv_rows(csv_text)
        assert rows == []


class TestDecodePayloadAlias:
    def test_decode_payload__동일_결과(self):
        payload = "한글".encode("utf-8")
        assert decode_payload(payload) == decode_csv_payload(payload)


class TestAnalyzeChatWithRows:
    def test_rows_파라미터_직접_전달(self):
        rows = [
            ("user1", "3/15😀"),
            ("user1", "3/16😀"),
        ]
        result = analyze_chat(rows=rows)
        assert "user1" in result
        assert result["user1"]["dates"] == {"3/15", "3/16"}

    def test_rows_파라미터_dual_모드(self):
        rows = [
            ("user1", "2/2 구약 😀"),
            ("user1", "2/3 신약 😀"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2"}
        assert result["user1"]["dates_new"] == {"2/3"}

    def test_csv_text와_rows_동시_전달시_rows_우선(self):
        rows = [("user1", "3/15😀")]
        result = analyze_chat(csv_text="invalid", rows=rows)
        assert "user1" in result


class TestBuildDualPreviewData:
    def test_구약_신약_분리된_헤더와_행(self):
        users = {
            "user1": {"dates_old": {"2/2", "2/4"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        old_h, old_r, new_h, new_r = build_dual_preview_data(users)
        assert old_h[:2] == ["이름", "이모티콘"]
        assert "2/2" in old_h
        assert "2/4" in old_h
        assert len(old_r) == 1
        assert old_r[0][0] == "user1"

        assert new_h[:2] == ["이름", "이모티콘"]
        assert "2/3" in new_h
        assert len(new_r) == 1

    def test_한쪽_트랙만_있는_사용자(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": set(), "emoji": "😀"},
            "user2": {"dates_old": set(), "dates_new": {"2/3"}, "emoji": "🔥"},
        }
        old_h, old_r, new_h, new_r = build_dual_preview_data(users)
        assert len(old_r) == 1
        assert old_r[0][0] == "user1"
        assert len(new_r) == 1
        assert new_r[0][0] == "user2"

    def test_빈_사용자(self):
        old_h, old_r, new_h, new_r = build_dual_preview_data({})
        assert old_h == ["이름", "이모티콘"]
        assert old_r == []
        assert new_h == ["이름", "이모티콘"]
        assert new_r == []

    def test_xlsx와_동일_데이터(self):
        from openpyxl import load_workbook

        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        old_h, old_r, new_h, new_r = build_dual_preview_data(users)
        result = build_output_xlsx(users, track_mode="dual")
        wb = load_workbook(io.BytesIO(result))
        ws_old = wb["구약 진도표"]
        ws_new = wb["신약 진도표"]

        # 구약 헤더 일치
        xlsx_old_h = [ws_old.cell(2, c).value for c in range(2, ws_old.max_column + 1)]
        assert old_h == xlsx_old_h

        # 신약 헤더 일치
        xlsx_new_h = [ws_new.cell(2, c).value for c in range(2, ws_new.max_column + 1)]
        assert new_h == xlsx_new_h


class TestAnalyzeChatMultilineMessage:
    """멀티라인 메시지를 줄별로 분리하여 처리하는 테스트."""

    def test_멀티라인_선행_틸드__줄별_트랙_분리_dual(self):
        """서조은 케이스: ~ 2/25 구약 🍚 / ~ 3/2 신약 🍚"""
        rows = [
            ("user1", "2/20 구약 신약 🍚"),
            ("user1", "~ 2/25 구약 🍚\n~ 3/2 신약 🍚"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "user1" in result
        # 구약: 2/20 + 2/21~2/25
        assert "2/20" in result["user1"]["dates_old"]
        assert "2/25" in result["user1"]["dates_old"]
        # 신약: 2/20 + 2/21~3/2
        assert "2/20" in result["user1"]["dates_new"]
        assert "3/2" in result["user1"]["dates_new"]

    def test_멀티라인_날짜별_트랙_분리_dual(self):
        """2/26 구약 🍚 / 3/3 신약 🍚"""
        rows = [
            ("user1", "2/26 구약 🍚\n3/3 신약 🍚"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "user1" in result
        assert "2/26" in result["user1"]["dates_old"]
        assert "2/26" not in result["user1"]["dates_new"]
        assert "3/3" in result["user1"]["dates_new"]
        assert "3/3" not in result["user1"]["dates_old"]

    def test_멀티라인_single_모드(self):
        rows = [
            ("user1", "3/6 😀\n3/7 😀"),
        ]
        result = analyze_chat(rows=rows, track_mode="single")
        assert "user1" in result
        assert "3/6" in result["user1"]["dates"]
        assert "3/7" in result["user1"]["dates"]

    def test_단일라인__변경_없음(self):
        rows = [
            ("user1", "3/6 구약 신약 😀"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "user1" in result
        assert "3/6" in result["user1"]["dates_old"]
        assert "3/6" in result["user1"]["dates_new"]

    def test_멀티라인_빈줄_포함(self):
        """2/28 구약 신약 (무표정) / / 3/2 구약 신약 (무표정)"""
        rows = [
            ("광천 강창우", "2/28 구약 신약  (무표정)\n\n3/2 구약 신약  (무표정)"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "강창우" in result
        assert "2/28" in result["강창우"]["dates_old"]
        assert "3/2" in result["강창우"]["dates_old"]


class TestAnalyzeChatEmojiChange:
    """사용자가 이모지를 변경한 경우 양쪽 모두 인식하는 테스트."""

    def test_이모지_변경__양쪽_카운트(self):
        """이창원 케이스: 🍯 → ⭐️ 변경 후에도 양쪽 다 인식."""
        rows = [
            ("user1", "2/2 구약 신약 🍯"),
            ("user1", "2/3 구약 신약 🍯"),
            ("user1", "2/4 구약 신약 ⭐️"),
            ("user1", "2/5 구약 신약 ⭐️"),
        ]
        result = analyze_chat(rows=rows, track_mode="dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2", "2/3", "2/4", "2/5"}
        assert result["user1"]["dates_new"] == {"2/2", "2/3", "2/4", "2/5"}

    def test_이모지_변경__대표_이모지는_최다(self):
        """대표 이모지는 가장 많이 사용한 것."""
        rows = [
            ("user1", "2/2 😀"),
            ("user1", "2/3 😀"),
            ("user1", "2/4 😀"),
            ("user1", "2/5 🔥"),
        ]
        result = analyze_chat(rows=rows)
        assert result["user1"]["emoji"] == "😀"

    def test_이모지_변경__single_모드(self):
        rows = [
            ("user1", "3/6 😀"),
            ("user1", "3/7 🔥"),
        ]
        result = analyze_chat(rows=rows)
        assert "user1" in result
        assert "3/6" in result["user1"]["dates"]
        assert "3/7" in result["user1"]["dates"]


class TestAnalyzeChatConsecutiveMessageWithoutEmoji:
    """같은 사용자가 연속으로 보낸 메시지에서 이모지 생략 시 허용 테스트."""

    def _make_csv(self, rows):
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in rows:
            writer.writerow(row)
        return output.getvalue()

    def test_연속_메시지_이모지_생략__날짜_카운트(self):
        """강창우 케이스: 이모지 포함 메시지 직후 이모지 없는 메시지도 카운트."""
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/6 구약 신약 😀"],
            ["2024-01-01", "user1", "3/7 구약"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "3/6" in result["user1"]["dates"]
        assert "3/7" in result["user1"]["dates"]

    def test_비연속_메시지_이모지_생략__카운트_안됨(self):
        """다른 사용자 메시지가 사이에 있으면 이모지 없는 메시지는 무시."""
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/6 구약 신약 😀"],
            ["2024-01-01", "user2", "3/6 구약 🔥"],
            ["2024-01-01", "user1", "3/7 구약"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert "3/6" in result["user1"]["dates"]
        assert "3/7" not in result["user1"]["dates"]

    def test_연속_메시지_이모지_생략__dual_모드(self):
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/6 구약 신약 😀"],
            ["2024-01-01", "user1", "3/7 구약"],
        ])
        result = analyze_chat(csv_text, track_mode="dual")
        assert "user1" in result
        assert "3/6" in result["user1"]["dates_old"]
        assert "3/6" in result["user1"]["dates_new"]
        assert "3/7" in result["user1"]["dates_old"]

    def test_이모지_없고_날짜도_없는_메시지__무시(self):
        """날짜 없는 메시지는 연속이더라도 무시."""
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/6 구약 신약 😀"],
            ["2024-01-01", "user1", "안녕하세요"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"3/6"}

    def test_연속_이모지_생략_후_다시_이모지_포함(self):
        """이모지 생략 메시지 후 다시 이모지 포함 메시지가 오는 경우."""
        csv_text = self._make_csv([
            ["날짜", "이름", "메시지"],
            ["2024-01-01", "user1", "3/6 구약 신약 😀"],
            ["2024-01-01", "user1", "3/7 구약"],
            ["2024-01-02", "user1", "3/9 구약 신약 😀"],
        ])
        result = analyze_chat(csv_text)
        assert "user1" in result
        assert result["user1"]["dates"] == {"3/6", "3/7", "3/9"}


class TestNormalizeUserName:
    def test_숫자_제거(self):
        assert normalize_user_name("김신영99") == "김신영"

    def test_영어_제거(self):
        assert normalize_user_name("홍길동ABC") == "홍길동"

    def test_공백_제거(self):
        assert normalize_user_name("홍 길 동") == "홍길동"

    def test_이모지_제거(self):
        assert normalize_user_name("🍯김신영") == "김신영"

    def test_키워드_제거__광천(self):
        assert normalize_user_name("광천유영훈") == "유영훈"

    def test_키워드_제거__누나(self):
        assert normalize_user_name("예슬누나") == "예슬"

    def test_키워드_제거__오빠(self):
        assert normalize_user_name("철수오빠") == "철수"

    def test_키워드_제거__언니(self):
        assert normalize_user_name("영희언니") == "영희"

    def test_복합__이모지_숫자_영어_공백_키워드(self):
        assert normalize_user_name("🍯 광천 유영훈 99 ABC") == "유영훈"

    def test_모두_제거되면__원본_반환(self):
        assert normalize_user_name("ABC 123") == "ABC 123"

    def test_형_공백포함_제거(self):
        assert normalize_user_name("철수 형") == "철수"

    def test_형_붙어있으면_유지(self):
        assert normalize_user_name("철수형") == "철수형"

    def test_이름에_형_포함__유지(self):
        assert normalize_user_name("김도형") == "김도형"

    def test_한글만__그대로(self):
        assert normalize_user_name("김철수") == "김철수"
