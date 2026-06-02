import datetime
import io
import json
import os
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from app.analyzer import build_output_xlsx
from app.schedule import BIBLE_PART_DATES, NT_PART_DATES
from app.style_constants import COL_PAD, ROW_PAD, apply_sheet_style
from app.merger import (
    _classify_education_users,
    _compute_dual_stats,
    _format_sheet_stats,
    _extract_date_from_filename,
    _extract_room_from_filename,
    _insert_stats_row,
    load_education_config as _load_education_config,
    build_merged_preview,
    build_merged_xlsx,
    merge_files,
    read_meta_from_xlsx,
    read_users_from_xlsx,
    resolve_alias,
    resolve_leader_override,
    select_latest_per_room,
)


class TestResolveAlias:
    def test_정확_일치__변환(self):
        aliases = {"태환": "김태환", "지혜": "홍지혜"}
        assert resolve_alias("태환", aliases) == "김태환"

    def test_정확_일치__별칭_등록된_이름_변환(self):
        aliases = {"조갑경": "조가빈", "영보조갑경": "조가빈"}
        assert resolve_alias("영보조갑경", aliases) == "조가빈"

    def test_매칭_없음__원래_이름_반환(self):
        aliases = {"태환": "김태환"}
        assert resolve_alias("김철수", aliases) == "김철수"

    def test_정확_일치_우선(self):
        aliases = {"태환": "김태환", "태": "다른사람"}
        assert resolve_alias("태환", aliases) == "김태환"

    def test_빈_별칭__원래_이름_반환(self):
        assert resolve_alias("김철수", {}) == "김철수"



    def test_정상_파일명__날짜시간_추출(self):
        assert _extract_date_from_filename("꿀성경_방장_20260210_1050_방이름.xlsx") == "20260210_1050"

    def test_방이름_없는_파일명__날짜시간_추출(self):
        assert _extract_date_from_filename("꿀성경_방장_20260210_1050.xlsx") == "20260210_1050"

    def test_패턴_불일치__None_반환(self):
        assert _extract_date_from_filename("기타파일.xlsx") is None

    def test_None__None_반환(self):
        assert _extract_date_from_filename(None) is None

    def test_빈_문자열__None_반환(self):
        assert _extract_date_from_filename("") is None


class TestExtractRoomFromFilename:
    def test_정상_파일명__방이름_추출(self):
        name = "꿀성경_방장_20260210_1050_2026 성경일독 part1.xlsx"
        assert _extract_room_from_filename(name) == "2026 성경일독 part1"

    def test_방이름에_언더스코어__재결합(self):
        name = "꿀성경_방장_20260210_1050_교육국_방.xlsx"
        assert _extract_room_from_filename(name) == "교육국_방"

    def test_패턴_불일치__파일명_그대로(self):
        name = "기타파일.xlsx"
        assert _extract_room_from_filename(name) == "기타파일.xlsx"

    def test_None__None_반환(self):
        assert _extract_room_from_filename(None) is None

    def test_빈_문자열__빈_문자열_반환(self):
        assert _extract_room_from_filename("") == ""

    def test_꿀성경_접두사_없음__파일명_그대로(self):
        name = "결과_방장_20260210_1050_방이름.xlsx"
        assert _extract_room_from_filename(name) == "결과_방장_20260210_1050_방이름.xlsx"

    def test_날짜_형식_불일치__파일명_그대로(self):
        name = "꿀성경_방장_2026_1050_방이름.xlsx"
        assert _extract_room_from_filename(name) == "꿀성경_방장_2026_1050_방이름.xlsx"


class TestSelectLatestPerRoom:
    def test_같은_방_여러_파일__최신만_선택(self):
        files = [
            {"id": "1", "name": "꿀성경_방장_20260210_1050_방1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"},
            {"id": "2", "name": "꿀성경_방장_20260211_1050_방1.xlsx", "modifiedTime": "2026-02-11T10:50:00Z"},
            {"id": "3", "name": "꿀성경_방장_20260210_1050_방2.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"},
        ]
        result = select_latest_per_room(files)
        assert len(result) == 2
        ids = {f["id"] for f in result}
        assert "2" in ids  # 방1의 최신
        assert "3" in ids  # 방2

    def test_방_하나__그대로(self):
        files = [
            {"id": "1", "name": "꿀성경_방장_20260210_1050_방1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"},
        ]
        result = select_latest_per_room(files)
        assert len(result) == 1

    def test_빈_리스트__빈_결과(self):
        assert select_latest_per_room([]) == []


class TestReadMetaFromXlsx:
    def test_메타_시트_있음__dict_반환(self):
        users = {"user1": {"dates": {"2/2"}, "emoji": "😀"}}
        meta = {"room_name": "테스트방", "track_mode": "single", "schedule_type": "bible", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        result = read_meta_from_xlsx(xlsx_bytes)
        assert result is not None
        assert result["room_name"] == "테스트방"
        assert result["track_mode"] == "single"
        assert result["schedule_type"] == "bible"
        assert result["leader"] == "방장"

    def test_메타_시트_없음__None_반환(self):
        users = {"user1": {"dates": {"2/2"}, "emoji": "😀"}}
        xlsx_bytes = build_output_xlsx(users, track_mode="single")

        result = read_meta_from_xlsx(xlsx_bytes)
        assert result is None

    def test_잘못된_바이트__None_반환(self):
        result = read_meta_from_xlsx(b"not an xlsx file")
        assert result is None


class TestReadUsersFromXlsx:
    def test_single_모드__사용자_데이터_추출(self):
        users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"},
            "user2": {"dates": {"2/2"}, "emoji": "🔥"},
        }
        xlsx_bytes = build_output_xlsx(users, track_mode="single")

        result = read_users_from_xlsx(xlsx_bytes, "single")
        assert "user1" in result
        assert result["user1"]["dates"] == {"2/2", "2/3"}
        assert result["user1"]["emoji"] == "😀"
        assert "user2" in result
        assert result["user2"]["dates"] == {"2/2"}

    def test_dual_모드__사용자_데이터_추출(self):
        users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
        }
        xlsx_bytes = build_output_xlsx(users, track_mode="dual")

        result = read_users_from_xlsx(xlsx_bytes, "dual")
        assert "user1" in result
        assert result["user1"]["dates_old"] == {"2/2"}
        assert result["user1"]["dates_new"] == {"2/3"}

    def test_빈_xlsx__빈_결과(self):
        users = {}
        xlsx_bytes = build_output_xlsx(users, track_mode="single")

        result = read_users_from_xlsx(xlsx_bytes, "single")
        assert result == {}

    def test_날짜형_헤더__월일_문자열로_정규화(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "꿀성경 진도표"
        ws.cell(2, 2, "이름")
        ws.cell(2, 3, "이모티콘")
        ws.cell(2, 4, datetime.datetime(2026, 2, 2))
        ws.cell(2, 5, "2026-02-03 00:00:00")
        ws.cell(3, 2, "user1")
        ws.cell(3, 3, "😀")
        ws.cell(3, 4, "O")
        ws.cell(3, 5, "O")

        buf = io.BytesIO()
        wb.save(buf)

        result = read_users_from_xlsx(buf.getvalue(), "single")
        assert result["user1"]["dates"] == {"2/2", "2/3"}


class TestClassifyEducationUsers:
    def test_정상_분류(self):
        users = {
            "김철수": {"dates": {"2/2"}, "emoji": "😀"},
            "지혜": {"dates": {"2/3"}, "emoji": "🔥"},
            "찬영": {"dates": {"2/4"}, "emoji": "🎉"},
            "지혁": {"dates": {"2/5"}, "emoji": "💀"},
        }
        config = {"nt_members": ["지혜", "찬영"], "excluded_members": ["지혁"]}

        result = _classify_education_users(users, config)
        assert "김철수" in result["bible"]
        assert "지혜" in result["nt"]
        assert "찬영" in result["nt"]
        assert "지혁" not in result["bible"]
        assert "지혁" not in result["nt"]

    def test_빈_설정__모두_성경일독(self):
        users = {
            "김철수": {"dates": {"2/2"}, "emoji": "😀"},
            "이영희": {"dates": {"2/3"}, "emoji": "🔥"},
        }
        config = {"nt_members": [], "excluded_members": []}

        result = _classify_education_users(users, config)
        assert len(result["bible"]) == 2
        assert len(result["nt"]) == 0

    def test_모두_제외__빈_결과(self):
        users = {
            "지혁": {"dates": {"2/2"}, "emoji": "😀"},
        }
        config = {"nt_members": [], "excluded_members": ["지혁"]}

        result = _classify_education_users(users, config)
        assert len(result["bible"]) == 0
        assert len(result["nt"]) == 0

    def test_부분_일치__닉네임에_키워드_포함(self):
        users = {
            "김철수": {"dates": {"2/2"}, "emoji": "😀"},
            "김지혜": {"dates": {"2/3"}, "emoji": "🔥"},
            "이찬영": {"dates": {"2/4"}, "emoji": "🎉"},
            "박지혁": {"dates": {"2/5"}, "emoji": "💀"},
        }
        config = {"nt_members": ["지혜", "찬영"], "excluded_members": ["지혁"]}

        result = _classify_education_users(users, config)
        assert "김철수" in result["bible"]
        assert "김지혜" in result["nt"]
        assert "이찬영" in result["nt"]
        assert "박지혁" not in result["bible"]
        assert "박지혁" not in result["nt"]


class TestBuildMergedXlsx:
    def test_양쪽_시트_생성(self):
        bible_users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {
            "user2": {"dates": {"2/4"}, "emoji": "🔥", "leader": "방장B"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, nt_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert "성경일독 진도표" in wb.sheetnames
        assert "신약일독 진도표" in wb.sheetnames
        assert "완독자" in wb.sheetnames
        assert "분석결과" in wb.sheetnames

    def test_담당_컬럼_포함(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {}
        xlsx_bytes = build_merged_xlsx(bible_users, nt_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5 = 데이터 (패딩 적용)
        assert ws.cell(4, 2).value == "담당"
        assert ws.cell(4, 3).value == "이름"
        assert ws.cell(4, 4).value == "이모티콘"
        assert ws.cell(5, 2).value == "방장A"
        assert ws.cell(5, 3).value == "user1"

    def test_담당별_정렬(self):
        bible_users = {
            "user_z": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장B"},
            "user_a": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5~ = 데이터 (방장A가 먼저)
        assert ws.cell(5, 2).value == "방장A"
        assert ws.cell(6, 2).value == "방장B"

    def test_빈_사용자__헤더만(self):
        xlsx_bytes = build_merged_xlsx({}, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]
        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5 = 데이터 없음
        assert ws.cell(4, 2).value == "담당"
        assert ws.cell(5, 2).value is None

    def test_담당_셀_병합_후_아래_테두리_두꺼운_선(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
            "user2": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장A"},
            "user3": {"dates": {"2/2"}, "emoji": "🌟", "leader": "방장B"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5-6 = 방장A, row 7 = 방장B
        leader_col = 2  # COL_PAD=1 → 담당 컬럼은 2

        # 병합된 셀(방장A, row 5-6)의 상단 셀에 medium 아래 테두리
        top_cell = ws.cell(5, leader_col)
        assert top_cell.border.bottom.style == "medium"

        # 병합 범위의 마지막 행(row 6)에도 medium 아래 테두리
        bottom_cell = ws.cell(6, leader_col)
        assert bottom_cell.border.bottom.style == "medium"

    def test_완독자_시트와_행_강조(self):
        bible_users = {
            "complete": {"dates": set(BIBLE_PART_DATES[0]), "emoji": "😀", "leader": "방장A"},
            "partial": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장A"},
        }
        nt_users = {
            "nt_complete": {"dates": set(NT_PART_DATES[0]), "emoji": "✨", "leader": "방장B"},
        }

        xlsx_bytes = build_merged_xlsx(bible_users, nt_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        ws_bible = wb["성경일독 진도표"]
        assert ws_bible.cell(5, 3).value == "complete"
        assert ws_bible.cell(5, 2).fill.start_color.rgb == "00EBF1F8"
        assert ws_bible.cell(5, 3).fill.start_color.rgb == "00CFE2F3"
        assert ws_bible.cell(5, 4).fill.start_color.rgb == "00CFE2F3"
        assert ws_bible.cell(5, 5).fill.start_color.rgb == "00CFE2F3"
        assert ws_bible.cell(6, 3).value == "partial"
        assert ws_bible.cell(6, 3).fill.start_color.rgb == "00EBF1F8"
        assert ws_bible.cell(6, 4).fill.start_color.rgb == "00000000"

        ws_complete = wb["완독자"]
        assert ws_complete.cell(4, 6).value is None
        assert ws_complete.cell(3, 2).fill.start_color.rgb == "00EBF1F8"
        assert ws_complete.cell(3, 3).fill.start_color.rgb == "00000000"
        assert ws_complete.cell(3, 4).fill.start_color.rgb == "00CFE2F3"
        assert ws_complete.cell(3, 5).fill.start_color.rgb == "00000000"
        completion_pairs = {
            (ws_complete.cell(row, 3).value, ws_complete.cell(row, 4).value)
            for row in range(3, ws_complete.max_row + 1)
        }
        assert ("성경일독", "complete") in completion_pairs
        assert ("신약일독", "nt_complete") in completion_pairs
        assert ("성경일독", "partial") not in completion_pairs

        ws_analysis = wb["분석결과"]
        assert ws_analysis.cell(2, 2).value == "분석결과"
        assert len(ws_analysis._charts) >= 3

    def test_분석결과__교육국_멤버만_이름_중복제거(self):
        bible_users = {
            "김태환": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": "방장A"},
            "김치훈": {"dates": {"2/2"}, "emoji": "🌿", "leader": "방장A"},
            "동명이인": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장A"},
        }
        nt_users = {
            "김태환": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장B"},
            "김치훈": {"dates": {"2/2"}, "emoji": "🌿", "leader": "방장B"},
            "동명이인": {"dates": {"2/2"}, "emoji": "✨", "leader": "방장B"},
        }

        xlsx_bytes = build_merged_xlsx(bible_users, nt_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["분석결과"]

        # row 5 = 요약 헤더, row 6 = 전체 요약. 김태환은 방장이라 1명,
        # 김치훈은 room_members의 일반 참여자라 2명, 동명이인도 2명으로 유지된다.
        assert ws.cell(6, 2).value == "전체"
        assert ws.cell(6, 3).value.startswith("=COUNTA(")

        ws_helper = wb["_분석계산"]
        assert ws_helper.sheet_state == "hidden"
        names = [ws_helper.cell(row, 4).value for row in range(2, ws_helper.max_row + 1)]
        assert names.count("김태환") == 1
        assert names.count("김치훈") == 2
        assert names.count("동명이인") == 2

    def test_분석결과__마지막_인증일은_오른쪽_끝_O를_참조(self):
        bible_users = {
            "김가람": {"dates": {"2/2", "5/13"}, "emoji": "🍫", "leader": "예슬"},
        }

        xlsx_bytes = build_merged_xlsx(bible_users, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_helper = wb["_분석계산"]
        row = next(
            row
            for row in range(2, ws_helper.max_row + 1)
            if ws_helper.cell(row, 4).value == "김가람"
        )

        formula = ws_helper.cell(row, 14).value
        assert "LOOKUP(" not in formula
        assert "MAX(FILTER(COLUMN('성경일독 진도표'!" in formula
        assert ws_helper.cell(row, 10).value == f"=N{row}"


class TestBuildMergedPreview:
    def test_양쪽_사용자_포함(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {
            "user2": {"dates": {"2/3"}, "emoji": "🔥", "leader": "방장B"},
        }
        headers, rows = build_merged_preview(bible_users, nt_users)

        assert "담당" in headers
        assert "트랙" in headers
        assert len(rows) == 2
        # 첫 행: 성경일독
        assert rows[0][3] == "성경일독"
        # 둘째 행: 신약일독
        assert rows[1][3] == "신약일독"

    def test_빈_사용자__행_없음(self):
        headers, rows = build_merged_preview({}, {})
        assert len(rows) == 0


class TestBuildMergedXlsxDualUsers:
    def test_dual_users_전달시_5개_시트_생성(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {
            "user2": {"dates": {"2/3"}, "emoji": "🔥", "leader": "방장B"},
        }
        dual_users = {
            "user3": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "🎉", "leader": "방장C"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, nt_users, dual_users=dual_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert "성경일독 진도표" in wb.sheetnames
        assert "신약일독 진도표" in wb.sheetnames
        assert "투트랙 진도표" in wb.sheetnames
        assert "완독자" in wb.sheetnames
        assert "분석결과" in wb.sheetnames

    def test_dual_users_None__완독자_분석결과_포함_4개_시트_유지(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {}
        xlsx_bytes = build_merged_xlsx(bible_users, nt_users, dual_users=None)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert "성경일독 진도표" in wb.sheetnames
        assert "신약일독 진도표" in wb.sheetnames
        assert "완독자" in wb.sheetnames
        assert "분석결과" in wb.sheetnames
        assert "투트랙 진도표" not in wb.sheetnames

    def test_dual_users_빈_dict__완독자_분석결과_포함_4개_시트_유지(self):
        xlsx_bytes = build_merged_xlsx({}, {}, dual_users={})
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert len(wb.sheetnames) == 5
        assert "투트랙 진도표" not in wb.sheetnames

    def test_투트랙_시트_구약_신약_행_분리(self):
        dual_users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx({}, {}, dual_users=dual_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["투트랙 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5-6 = 데이터
        assert ws.cell(4, 2).value == "담당"
        assert ws.cell(4, 3).value == "이름"
        assert ws.cell(4, 4).value == "이모티콘"
        assert ws.cell(4, 5).value == "트랙"
        # 구약 행
        assert ws.cell(5, 2).value == "방장A"
        assert ws.cell(5, 3).value == "user1"
        assert ws.cell(5, 5).value == "구약"
        # 신약 행
        assert ws.cell(6, 3).value == "user1"
        assert ws.cell(6, 5).value == "신약"

    def test_투트랙_시트_한쪽_트랙만_있는_사용자(self):
        dual_users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": set(), "emoji": "😀", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx({}, {}, dual_users=dual_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["투트랙 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5 = 데이터
        # 구약 행만 존재
        assert ws.cell(5, 3).value == "user1"
        assert ws.cell(5, 5).value == "구약"
        # 신약 행 없음
        assert ws.cell(6, 3).value is None

    def test_투트랙_완독자_시트와_트랙별_강조(self):
        dual_users = {
            "both": {
                "dates_old": set(BIBLE_PART_DATES[0]),
                "dates_new": set(NT_PART_DATES[0]),
                "emoji": "😀",
                "leader": "방장A",
            },
            "old_only": {
                "dates_old": set(BIBLE_PART_DATES[0]),
                "dates_new": {"2/2"},
                "emoji": "🔥",
                "leader": "방장A",
            },
        }
        xlsx_bytes = build_merged_xlsx({}, {}, dual_users=dual_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        ws_dual = wb["투트랙 진도표"]
        assert ws_dual.cell(5, 3).value == "both"
        assert ws_dual.cell(5, 3).fill.start_color.rgb == "00CFE2F3"
        assert ws_dual.cell(5, 4).fill.start_color.rgb == "00CFE2F3"
        assert ws_dual.cell(5, 5).fill.start_color.rgb == "00CFE2F3"
        assert ws_dual.cell(6, 3).value == "both"
        assert ws_dual.cell(6, 3).fill.start_color.rgb == "00CFE2F3"
        assert ws_dual.cell(7, 3).value == "old_only"
        assert ws_dual.cell(7, 3).fill.start_color.rgb == "00CFE2F3"
        assert ws_dual.cell(8, 3).value == "old_only"
        assert ws_dual.cell(8, 3).fill.start_color.rgb == "00EBF1F8"
        assert ws_dual.cell(8, 4).fill.start_color.rgb == "00000000"

        ws_complete = wb["완독자"]
        assert ws_complete.cell(4, 6).value is None
        assert ws_complete.cell(3, 2).fill.start_color.rgb == "00EBF1F8"
        assert ws_complete.cell(3, 3).fill.start_color.rgb == "00000000"
        assert ws_complete.cell(3, 4).fill.start_color.rgb == "00CFE2F3"
        assert ws_complete.cell(3, 5).fill.start_color.rgb == "00000000"
        completion_pairs = {
            (ws_complete.cell(row, 3).value, ws_complete.cell(row, 4).value)
            for row in range(3, ws_complete.max_row + 1)
        }
        assert ("투트랙", "both") in completion_pairs
        assert ("투트랙(구약)", "both") not in completion_pairs
        assert ("투트랙(신약)", "both") not in completion_pairs
        assert ("투트랙(둘 다)", "both") not in completion_pairs
        assert ("투트랙", "old_only") not in completion_pairs
        assert ("투트랙(구약)", "old_only") not in completion_pairs
        assert ("투트랙(신약)", "old_only") not in completion_pairs


class TestBuildMergedPreviewDualUsers:
    def test_dual_users_전달시_투트랙_구약신약_트랙_포함(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {}
        dual_users = {
            "user2": {"dates_old": {"2/3"}, "dates_new": {"2/4"}, "emoji": "🔥", "leader": "방장C"},
        }
        headers, rows = build_merged_preview(bible_users, nt_users, dual_users=dual_users)

        assert "트랙" in headers
        tracks = [row[3] for row in rows]
        assert "성경일독" in tracks
        assert "투트랙(구약)" in tracks
        assert "투트랙(신약)" in tracks

    def test_dual_users_None__기존_동작(self):
        bible_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        nt_users = {
            "user2": {"dates": {"2/3"}, "emoji": "🔥", "leader": "방장B"},
        }
        headers, rows = build_merged_preview(bible_users, nt_users, dual_users=None)

        tracks = [row[3] for row in rows]
        assert "투트랙(구약)" not in tracks
        assert "투트랙(신약)" not in tracks
        assert len(rows) == 2


class TestMergeFiles:
    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_성경일독_파일_통합(self, mock_list, mock_download):
        # 성경일독 XLSX 생성
        users = {"user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"}}
        meta = {"room_name": "part1", "track_mode": "single", "schedule_type": "bible", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_part1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "꿀성경_방장_20260210_1050_part1.xlsx"}

        result = merge_files()
        assert result["success"] is True
        assert "user1" in result["bible_users"]
        assert result["bible_users"]["user1"]["dates"] == {"2/2", "2/3"}
        assert len(result["nt_users"]) == 0
        assert result["oldest_file_date"] == "20260210_1050"

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_신약일독_파일_통합(self, mock_list, mock_download):
        users = {"user1": {"dates": {"2/2"}, "emoji": "😀"}}
        meta = {"room_name": "nt1", "track_mode": "single", "schedule_type": "nt", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_nt1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        result = merge_files()
        assert result["success"] is True
        assert "user1" in result["nt_users"]
        assert len(result["bible_users"]) == 0

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_듀얼_파일_split_모드__양쪽_분배(self, mock_list, mock_download):
        users = {"user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"}}
        meta = {"room_name": "dual방", "track_mode": "dual", "schedule_type": "dual", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="dual", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_dual방.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        result = merge_files(dual_mode="split")
        assert result["success"] is True
        assert "user1" in result["bible_users"]
        assert result["bible_users"]["user1"]["dates"] == {"2/2"}
        assert "user1" in result["nt_users"]
        assert result["nt_users"]["user1"]["dates"] == {"2/3"}
        assert len(result["dual_users"]) == 0

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_듀얼_파일_separate_모드__별도_시트_분리(self, mock_list, mock_download):
        users = {"user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"}}
        meta = {"room_name": "dual방", "track_mode": "dual", "schedule_type": "dual", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="dual", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_dual방.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        result = merge_files(dual_mode="separate")
        assert result["success"] is True
        assert len(result["bible_users"]) == 0
        assert len(result["nt_users"]) == 0
        assert "user1" in result["dual_users"]
        assert result["dual_users"]["user1"]["dates_old"] == {"2/2"}
        assert result["dual_users"]["user1"]["dates_new"] == {"2/3"}

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_교육국_파일_분류(self, mock_list, mock_download):
        users = {
            "김철수": {"dates": {"2/2"}, "emoji": "😀"},
            "홍지혜": {"dates": {"2/3"}, "emoji": "🔥"},
            "박지혁": {"dates": {"2/4"}, "emoji": "💀"},
        }
        meta = {"room_name": "교육국", "track_mode": "single", "schedule_type": "education", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_교육국.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        result = merge_files()
        assert result["success"] is True
        assert "김철수" in result["bible_users"]
        assert "홍지혜" in result["nt_users"]
        assert "박지혁" not in result["bible_users"]
        assert "박지혁" not in result["nt_users"]

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_중복_사용자_날짜_합집합(self, mock_list, mock_download):
        # 두 방에 같은 사용자가 있는 경우
        users1 = {"user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"}}
        meta1 = {"room_name": "방1", "track_mode": "single", "schedule_type": "bible", "leader": "방장A"}
        xlsx1 = build_output_xlsx(users1, track_mode="single", meta=meta1)

        users2 = {"user1": {"dates": {"2/3", "2/4"}, "emoji": "😀"}}
        meta2 = {"room_name": "방2", "track_mode": "single", "schedule_type": "bible", "leader": "방장B"}
        xlsx2 = build_output_xlsx(users2, track_mode="single", meta=meta2)

        mock_list.return_value = {
            "success": True,
            "files": [
                {"id": "1", "name": "꿀성경_방장A_20260209_0900_방1.xlsx", "modifiedTime": "2026-02-09T09:00:00Z"},
                {"id": "2", "name": "꿀성경_방장B_20260210_1050_방2.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"},
            ],
        }
        mock_download.side_effect = [
            {"success": True, "data": xlsx1, "name": "test1.xlsx"},
            {"success": True, "data": xlsx2, "name": "test2.xlsx"},
        ]

        result = merge_files()
        assert result["success"] is True
        assert result["bible_users"]["user1"]["dates"] == {"2/2", "2/3", "2/4"}
        assert result["oldest_file_date"] == "20260209_0900"

    @patch("app.merger.list_drive_files")
    def test_Drive_실패__에러_반환(self, mock_list):
        mock_list.return_value = {"success": False, "message": "API 오류"}
        result = merge_files()
        assert result["success"] is False

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_메타데이터_없는_파일__스킵(self, mock_list, mock_download):
        # 메타 없는 XLSX
        users = {"user1": {"dates": {"2/2"}, "emoji": "😀"}}
        xlsx_bytes = build_output_xlsx(users, track_mode="single")  # meta 미전달

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "old_file.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "old_file.xlsx"}

        result = merge_files()
        assert result["success"] is True
        assert len(result["skipped_files"]) == 1
        assert "메타데이터 없음" in result["skipped_files"][0]["reason"]

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_듀얼_separate_모드__dual_excluded_members_제외(self, mock_list, mock_download):
        users = {
            "이희준": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀"},
            "김철수": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "🔥"},
        }
        meta = {"room_name": "dual방", "track_mode": "dual", "schedule_type": "dual", "leader": "방장"}
        xlsx_bytes = build_output_xlsx(users, track_mode="dual", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_방장_20260210_1050_dual방.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        result = merge_files(dual_mode="separate")
        assert result["success"] is True
        assert "이희준" not in result["dual_users"]
        assert "김철수" in result["dual_users"]



class TestFormatSheetStats:
    def test_전원_완독(self):
        users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": ""},
            "user2": {"dates": {"2/2", "2/3"}, "emoji": "🔥", "leader": ""},
        }
        result = _format_sheet_stats(users, ["2/2", "2/3"])
        assert "진행: 2일" in result
        assert "참여: 2명" in result
        assert "완독: 2명 (100%)" in result

    def test_일부_완독(self):
        users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": ""},
            "user2": {"dates": {"2/2"}, "emoji": "🔥", "leader": ""},
        }
        result = _format_sheet_stats(users, ["2/2", "2/3"])
        assert "완독: 1명 (50%)" in result

    def test_완독_0명(self):
        users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": ""},
        }
        result = _format_sheet_stats(users, ["2/2", "2/3"])
        assert "완독: 0명 (0%)" in result

    def test_날짜_없는_멤버__전체_인원에_포함(self):
        users = {
            "user1": {"dates": set(), "emoji": "😀", "leader": ""},
        }
        result = _format_sheet_stats(users, ["2/2"])
        assert "참여: 1명" in result
        assert "완독: 0명 (0%)" in result


class TestComputeDualStats:
    def test_양쪽_완독(self):
        dual_users = {
            "user1": {"dates_old": {"2/2", "2/3"}, "dates_new": {"2/4", "2/5"}, "emoji": "😀"},
        }
        result = _compute_dual_stats(dual_users, ["2/2", "2/3", "2/4", "2/5"])
        assert "참여: 1명" in result
        assert "완독: 1명" in result

    def test_구약만_완독__완독_아님(self):
        # user1은 구약 다 했지만 신약 일부 누락, user2는 양쪽 다 했음
        dual_users = {
            "user1": {"dates_old": {"2/2", "2/3"}, "dates_new": {"2/4"}, "emoji": "😀"},
            "user2": {"dates_old": {"2/2", "2/3"}, "dates_new": {"2/4", "2/5"}, "emoji": "🔥"},
        }
        # all_old={2/2,2/3}, all_new={2/4,2/5}
        # user1: old ✓, new {2/4} < {2/4,2/5} ✗ → 미완독
        # user2: old ✓, new ✓ → 완독
        result = _compute_dual_stats(dual_users, ["2/2", "2/3", "2/4", "2/5"])
        assert "완독: 1명" in result

    def test_신약만_완독__완독_아님(self):
        # user1은 신약 다 했지만 구약 일부 누락, user2는 양쪽 다 했음
        dual_users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/4", "2/5"}, "emoji": "😀"},
            "user2": {"dates_old": {"2/2", "2/3"}, "dates_new": {"2/4", "2/5"}, "emoji": "🔥"},
        }
        # all_old={2/2,2/3}, all_new={2/4,2/5}
        # user1: old {2/2} < {2/2,2/3} ✗ → 미완독
        # user2: old ✓, new ✓ → 완독
        result = _compute_dual_stats(dual_users, ["2/2", "2/3", "2/4", "2/5"])
        assert "완독: 1명" in result

    def test_토요일_제외_구약_완독(self):
        # 2/7은 토요일 → 구약 expected에서 제외
        dual_users = {
            "user1": {
                "dates_old": {"2/2", "2/3"},  # 2/7(토) 없어도 완독
                "dates_new": {"2/4", "2/5"},
                "emoji": "😀",
            },
        }
        # 전체 날짜에 2/7 포함하지만, 구약 expected에서 토요일 제외
        # all_old = {2/2, 2/3}, all_new = {2/4, 2/5}
        # old_expected = {2/2, 2/3} (토요일 없음)
        result = _compute_dual_stats(dual_users, ["2/2", "2/3", "2/4", "2/5"])
        assert "완독: 1명" in result

    def test_토요일_포함_구약__토요일도_기대에_포함(self):
        # 구약은 월~토 읽기이므로 토요일도 완독 기준에 포함
        dual_users = {
            "user1": {
                "dates_old": {"2/2", "2/3"},
                "dates_new": {"2/4", "2/5"},
                "emoji": "😀",
            },
            "user2": {
                "dates_old": {"2/2", "2/3", "2/7"},  # 2/7 토요일 포함
                "dates_new": {"2/4", "2/5"},
                "emoji": "🔥",
            },
        }
        result = _compute_dual_stats(dual_users, ["2/2", "2/3", "2/4", "2/5", "2/7"])
        # old_expected = {2/2, 2/3, 2/7} (토요일 포함), new_expected = {2/4, 2/5}
        # user1: old {2/2, 2/3} < {2/2, 2/3, 2/7} → 미완독
        # user2: old {2/2, 2/3, 2/7} ✓, new {2/4, 2/5} ✓ → 완독
        assert "참여: 2명" in result
        assert "완독: 1명" in result


class TestInsertStatsRow:
    def test_삽입_후_row3에_통계_텍스트(self):
        wb = Workbook()
        ws = wb.active
        headers = ["담당", "이름", "이모티콘", "2/2"]
        rows = [["방장", "user1", "😀", "O"]]
        apply_sheet_style(ws, headers, rows, leader_col=1, title="테스트 타이틀")

        _insert_stats_row(ws, "진행: 1일 | 참여: 1명 | 완독: 1명 (100%)", len(headers))

        # stats row는 row 3에 삽입됨
        stats_row = 2 + ROW_PAD  # = 3
        assert ws.cell(stats_row, 1 + COL_PAD).value == "진행: 1일 | 참여: 1명 | 완독: 1명 (100%)"

    def test_freeze_panes_변경(self):
        wb = Workbook()
        ws = wb.active
        headers = ["담당", "이름", "이모티콘"]
        rows = []
        apply_sheet_style(ws, headers, rows, leader_col=1, title="테스트")

        assert ws.freeze_panes == "B4"
        _insert_stats_row(ws, "통계", len(headers))
        assert ws.freeze_panes == "B5"


class TestBuildMergedSheetStats:
    def test_성경일독_시트_통계_행_포함(self):
        bible_users = {
            "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": "방장A"},
            "user2": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5~ = 데이터
        stats_cell = ws.cell(3, 2)  # COL_PAD=1 → column 2
        assert stats_cell.value is not None
        assert "진행:" in stats_cell.value
        assert "참여:" in stats_cell.value
        assert "완독:" in stats_cell.value

    def test_투트랙_시트_통계_행_포함(self):
        dual_users = {
            "user1": {"dates_old": {"2/2"}, "dates_new": {"2/3"}, "emoji": "😀", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx({}, {}, dual_users=dual_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["투트랙 진도표"]

        stats_cell = ws.cell(3, 2)
        assert stats_cell.value is not None
        assert "진행:" in stats_cell.value

    def test_신약일독_시트_통계_행_포함(self):
        nt_users = {
            "user1": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx({}, nt_users)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["신약일독 진도표"]

        stats_cell = ws.cell(3, 2)
        assert stats_cell.value is not None
        assert "진행:" in stats_cell.value


class TestMergeFilesRoomMembers:
    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_room_members_누락_멤버_통합_결과에_포함(self, mock_list, mock_download):
        """room_members에 등록된 멤버 중 XLSX에 없는 인원이 빈 날짜로 통합 결과에 포함된다."""
        users = {"참여자A": {"dates": {"2/2"}, "emoji": "😀"}}
        meta = {"room_name": "방1", "track_mode": "single", "schedule_type": "bible", "leader": "태환"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_태환_20260210_1050_방1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        edu_config = {
            "nt_members": [],
            "excluded_members": [],
            "name_aliases": {"태환": "김태환"},
            "room_members": {
                "김태환": ["참여자A", "미참여자B"],
            },
        }

        with patch("app.merger.load_education_config", return_value=edu_config):
            result = merge_files()

        assert result["success"] is True
        assert "참여자A" in result["bible_users"]
        assert "미참여자B" in result["bible_users"]
        assert result["bible_users"]["참여자A"]["dates"] == {"2/2"}
        assert result["bible_users"]["미참여자B"]["dates"] == set()

    @patch("app.merger.download_drive_file")
    @patch("app.merger.list_drive_files")
    def test_room_members_excluded_멤버는_재추가되지_않음(self, mock_list, mock_download):
        """room_members에 등록되어 있어도 excluded_members인 경우 통합 결과에서 제외된다."""
        users = {"참여자A": {"dates": {"2/2"}, "emoji": "😀"}}
        meta = {"room_name": "방1", "track_mode": "single", "schedule_type": "bible", "leader": "태환"}
        xlsx_bytes = build_output_xlsx(users, track_mode="single", meta=meta)

        mock_list.return_value = {
            "success": True,
            "files": [{"id": "1", "name": "꿀성경_태환_20260210_1050_방1.xlsx", "modifiedTime": "2026-02-10T10:50:00Z"}],
        }
        mock_download.return_value = {"success": True, "data": xlsx_bytes, "name": "test.xlsx"}

        edu_config = {
            "nt_members": [],
            "excluded_members": ["제외자"],
            "name_aliases": {"태환": "김태환"},
            "room_members": {
                "김태환": ["참여자A", "제외자C"],
            },
        }

        with patch("app.merger.load_education_config", return_value=edu_config):
            result = merge_files()

        assert result["success"] is True
        assert "참여자A" in result["bible_users"]
        # "제외자C"는 "제외자" 키워드 포함 → excluded_members에 의해 제거
        assert "제외자C" not in result["bible_users"]


class TestResolveLeaderOverride:
    def test_매칭_조건_충족__actual_반환(self):
        users = {"강민정": {"dates": set()}, "김태현": {"dates": set()}, "홍길동": {"dates": set()}}
        overrides = [{"detected": "원예진", "markers": ["강민정", "김태현"], "actual": "이찬영"}]
        assert resolve_leader_override("원예진", users, overrides) == "이찬영"

    def test_markers_일부만_존재__오버라이드_안함(self):
        users = {"강민정": {"dates": set()}, "홍길동": {"dates": set()}}
        overrides = [{"detected": "원예진", "markers": ["강민정", "김태현"], "actual": "이찬영"}]
        assert resolve_leader_override("원예진", users, overrides) == "원예진"

    def test_detected_불일치__원래_leader_반환(self):
        users = {"강민정": {"dates": set()}, "김태현": {"dates": set()}}
        overrides = [{"detected": "원예진", "markers": ["강민정", "김태현"], "actual": "이찬영"}]
        assert resolve_leader_override("김태환", users, overrides) == "김태환"

    def test_overrides_빈_리스트__원래_leader_반환(self):
        users = {"강민정": {"dates": set()}}
        assert resolve_leader_override("원예진", users, []) == "원예진"


class TestBuildMergedSheetEmptyDates:
    def test_빈_날짜_사용자_통합_시트에_포함(self):
        """dates가 빈 사용자도 통합 시트에 행으로 포함된다."""
        bible_users = {
            "참여자A": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
            "미참여자B": {"dates": set(), "emoji": "", "leader": "방장A"},
        }
        xlsx_bytes = build_merged_xlsx(bible_users, {})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["성경일독 진도표"]

        # row 2 = 타이틀, row 3 = 통계, row 4 = 헤더, row 5~ = 데이터
        user_names = []
        for r in range(5, ws.max_row + 1):
            val = ws.cell(r, 3).value  # 이름 컬럼 (COL_PAD=1)
            if val:
                user_names.append(val)
        assert "참여자A" in user_names
        assert "미참여자B" in user_names

    def test_빈_날짜_사용자_미리보기에_포함(self):
        """dates가 빈 사용자도 미리보기에 행으로 포함된다."""
        bible_users = {
            "참여자A": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장A"},
            "미참여자B": {"dates": set(), "emoji": "", "leader": "방장A"},
        }
        headers, rows = build_merged_preview(bible_users, {})
        user_names = [row[1] for row in rows]
        assert "참여자A" in user_names
        assert "미참여자B" in user_names
