from openpyxl import Workbook

from app.analytics import (
    AnalysisRecord,
    GROUP_ALL,
    GROUP_BIBLE,
    GROUP_DUAL,
    GROUP_NT,
    add_analysis_sheet,
    activity_trend,
    build_merged_analysis_records,
    build_output_analysis_records,
    dedupe_record_count,
    dropout_distribution,
    dropout_week_distribution,
    dual_record_count,
    progress_distribution,
    summarize_records,
)
from app.schedule import BIBLE_PART_DATES, NT_PART_DATES


def test_single_성경일독_분석_레코드__완독_하차_미시작_분류():
    partial_dates = {"2/2", "2/3"}
    users = {
        "complete": {"dates": set(BIBLE_PART_DATES[0]), "emoji": "😀"},
        "partial": {"dates": partial_dates, "emoji": "🔥"},
        "none": {"dates": set(), "emoji": ""},
    }

    records = build_output_analysis_records(users, meta={"schedule_type": "bible", "part": 1})
    by_name = {record.name: record for record in records}

    assert by_name["complete"].complete is True
    assert by_name["complete"].status == "완독"
    assert by_name["partial"].status == "하차 추정"
    assert by_name["partial"].last_date == "2/3"
    assert "부근" in by_name["partial"].last_position
    assert by_name["none"].status == "미시작"


def test_dual_분석_레코드__투트랙을_한명으로_계산():
    users = {
        "dual": {
            "dates_old": {"2/2", "2/3"},
            "dates_new": {"2/2"},
            "emoji": "😀",
        },
    }

    records = build_output_analysis_records(users, track_mode="dual", meta={"part": 1})
    assert len(records) == 1
    record = records[0]
    assert record.group == GROUP_DUAL
    assert record.read_count == 3
    assert record.expected_count == len(BIBLE_PART_DATES[0]) + len(NT_PART_DATES[0])
    assert record.complete is False
    assert record.status == "하차 추정"

    summary = {row["group"]: row for row in summarize_records(records)}
    assert dual_record_count(records) == 1
    assert summary[GROUP_ALL]["total"] == 1
    assert summary[GROUP_DUAL]["total"] == 1


def test_merged_분석_레코드__전체와_3개_그룹_요약():
    bible_users = {
        "bible": {"dates": set(BIBLE_PART_DATES[0]), "emoji": "😀", "leader": "방장A"},
    }
    nt_users = {
        "nt": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장B"},
    }
    dual_users = {
        "dual": {"dates_old": {"2/2"}, "dates_new": {"2/2"}, "emoji": "✨", "leader": "방장C"},
    }

    records = build_merged_analysis_records(bible_users, nt_users, dual_users, part=1)
    summary = {row["group"]: row for row in summarize_records(records)}

    assert summary[GROUP_ALL]["total"] == 3
    assert summary[GROUP_BIBLE]["complete"] == 1
    assert summary[GROUP_DUAL]["total"] == 1
    assert summary[GROUP_DUAL]["dropout"] == 1


def test_merged_분석_레코드__교육국_dedupe_대상_이름만_중복_카운트하지_않음():
    bible_users = {
        "담당자": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": "방장A"},
    }
    nt_users = {
        "담당자": {"dates": {"2/2"}, "emoji": "😀", "leader": "방장B"},
    }

    records = build_merged_analysis_records(
        bible_users,
        nt_users,
        {},
        part=1,
        dedupe_names={"담당자"},
    )
    summary = {row["group"]: row for row in summarize_records(records)}

    assert dedupe_record_count(records, dedupe_names={"담당자"}) == 1
    assert len(records) == 1
    assert summary[GROUP_ALL]["total"] == 1


def test_merged_분석_레코드__일반_동명이인은_각각_카운트():
    bible_users = {
        "동명이인": {"dates": {"2/2", "2/3"}, "emoji": "😀", "leader": "방장A"},
    }
    nt_users = {
        "동명이인": {"dates": {"2/2"}, "emoji": "🔥", "leader": "방장B"},
    }

    records = build_merged_analysis_records(bible_users, nt_users, {}, part=1)
    summary = {row["group"]: row for row in summarize_records(records)}

    assert len(records) == 2
    assert summary[GROUP_ALL]["total"] == 2
    assert summary[GROUP_BIBLE]["total"] == 1
    assert summary[GROUP_NT]["total"] == 1


def test_진행률_분포와_하차_분포():
    users = {
        "complete": {"dates": set(BIBLE_PART_DATES[0]), "emoji": "😀"},
        "partial": {"dates": {"2/2", "2/3"}, "emoji": "🔥"},
        "none": {"dates": set(), "emoji": ""},
    }
    records = build_output_analysis_records(users, meta={"schedule_type": "bible", "part": 1})

    progress = {row["bucket"]: row for row in progress_distribution(records)}
    assert progress["0%"][GROUP_ALL] == 1
    assert progress["1~25%"][GROUP_ALL] == 1
    assert progress["100%"][GROUP_ALL] == 1

    dropout = dropout_distribution(records)
    assert len(dropout) == 1
    assert dropout[0]["week"] == "2/2~2/8"
    assert dropout[0]["date"] == "2/3"
    assert "2/2~2/8 |" in dropout[0]["label"]
    assert dropout[0]["count"] == 1
    assert dropout[0]["names"] == "partial"


def test_하차_분포__같은_주는_하나로_묶음():
    users = {
        "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"},
        "user2": {"dates": {"2/2", "2/4"}, "emoji": "🔥"},
    }
    records = build_output_analysis_records(users, meta={"schedule_type": "bible", "part": 1})

    dropout = dropout_distribution(records)
    assert len(dropout) == 1
    assert dropout[0]["week"] == "2/2~2/8"
    assert dropout[0]["date"] == "2/3, 2/4"
    assert dropout[0]["count"] == 2
    assert dropout[0]["names"] == "user1, user2"


def test_하차_주_분포__같은_주_여러_위치를_하나로_묶음():
    bible_users = {
        "bible": {"dates": {"2/2"}, "emoji": "😀"},
    }
    nt_users = {
        "nt": {"dates": {"2/2"}, "emoji": "🔥"},
    }
    records = build_merged_analysis_records(bible_users, nt_users, {}, part=1)

    dropout = dropout_week_distribution(records)
    assert len(dropout) == 1
    assert dropout[0]["week"] == "2/2~2/8"
    assert dropout[0]["date"] == "2/2"
    assert dropout[0]["count"] == 2
    assert "창세기" in dropout[0]["position"]
    assert "마태복음" in dropout[0]["position"]
    assert "부근" not in dropout[0]["position"]
    assert dropout[0]["bible"] == 1
    assert dropout[0]["nt"] == 1
    assert dropout[0]["dual"] == 0


def test_하차_주_분포__트랙_접두어와_중복_위치를_정리():
    records = [
        AnalysisRecord(
            group=GROUP_BIBLE,
            name="bible",
            emoji="",
            leader="",
            read_count=1,
            expected_count=10,
            complete=False,
            last_date="3/16",
            last_track=GROUP_BIBLE,
            last_position="사사기 부근",
            status="하차 추정",
            activity_dates=frozenset({"3/16"}),
        ),
        AnalysisRecord(
            group=GROUP_DUAL,
            name="dual",
            emoji="",
            leader="",
            read_count=2,
            expected_count=20,
            complete=False,
            last_date="3/16",
            last_track="구약/신약",
            last_position="구약: 사사기 부근 / 신약: 룻기 부근 / 구약: 사사기 부근 / 신약: 사사기 부근",
            status="하차 추정",
            activity_dates=frozenset({"3/16"}),
        ),
    ]

    dropout = dropout_week_distribution(records)

    assert len(dropout) == 1
    assert dropout[0]["position"] == "사사기 / 룻기"


def test_투트랙_신약_진행위치는_신약_권명으로_계산():
    dual_users = {
        "dual": {
            "dates_old": {"3/16"},
            "dates_new": {"3/16"},
            "emoji": "",
        }
    }

    records = build_merged_analysis_records({}, {}, dual_users, part=1)

    assert records[0].last_position == "구약: 사사기 부근 / 신약: 마가복음 부근"


def test_날짜별_인증_추이__사용자별_일자_집계():
    users = {
        "user1": {"dates": {"2/2", "2/3"}, "emoji": "😀"},
        "user2": {"dates": {"2/3"}, "emoji": "🔥"},
    }
    records = build_output_analysis_records(users, meta={"schedule_type": "bible", "part": 1})

    trend = {row["date"]: row["count"] for row in activity_trend(records)}
    assert trend == {"2/2": 1, "2/3": 2}


def test_add_analysis_sheet__표와_차트_생성():
    users = {
        "complete": {"dates": set(BIBLE_PART_DATES[0]), "emoji": "😀"},
        "partial": {"dates": {"2/2", "2/3"}, "emoji": "🔥"},
    }
    records = build_output_analysis_records(users, meta={"schedule_type": "bible", "part": 1})

    wb = Workbook()
    ws = add_analysis_sheet(wb, records)

    assert ws.title == "분석결과"
    assert ws.cell(2, 2).value == "분석결과"
    assert ws.cell(5, 2).value == "그룹"
    assert ws.cell(6, 2).value == "전체"
    assert ws.cell(20, 2).value == "합"
    assert ws.cell(24, 2).value == "주차"
    assert ws.cell(24, 3).value == "하차 주"
    assert ws.cell(24, 4).value == "진행 위치"
    assert ws.cell(24, 5).value == "성경일독"
    assert ws.cell(24, 6).value == "신약일독"
    assert ws.cell(24, 7).value == "투트랙"
    assert ws.cell(24, 8).value == "합"
    assert "마지막 인증일" not in [ws.cell(24, col).value for col in range(2, 9)]
    assert ws.cell(25, 2).value == "1주차"
    assert ws.cell(25, 3).value == "2/2~2/8"
    assert "|" not in ws.cell(25, 3).value
    assert ws.cell(25, 4).value == "창세기"
    assert "부근" not in ws.cell(25, 4).value
    assert ws.cell(25, 4).alignment.wrap_text is False
    assert ws.cell(25, 4).alignment.horizontal == "left"
    assert ws.cell(25, 5).value == 1
    assert ws.cell(25, 8).value == 1
    assert ws.cell(26, 2).value == "합"
    assert ws.cell(26, 5).value == 1
    assert ws.cell(26, 8).value == 1
    assert ws.column_dimensions["D"].width >= 40
    assert len(ws._charts) >= 3
    assert type(ws._charts[0]).__name__ == "LineChart"
    assert len(ws._charts[0].series) == 3
